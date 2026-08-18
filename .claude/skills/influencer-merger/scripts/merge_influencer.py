# -*- coding: utf-8 -*-
"""
유튜브 인플루언서 리스트 -> '유튜브 인플루언서 관리'.메일발송현황 시트 병합 (다중 타겟)
- 소스 폴더(공유 스프레드 폴더)의 유튜브_인플루언서_리스트*.csv/xlsx 전체를 mtime 순으로 처리
- 타겟 2개에 동일 소스를 각각 병합:
    1) 마스터 : C:\\Users\\user\\비서\\스케줄\\0. 유튜브 인플루언서 관리*.xlsx
    2) 공유   : (소스 폴더)\\1. 유튜브 인플루언서 관리_공유_*.xlsx
- 각 타겟은 자기 시트 기준으로 (채널명, 링크) 중복 자동 제거
- 각 타겟 백업 -> append -> 행 수 검증 (타겟별 독립)
- 소스 파일 이동은 "모든 타겟 정상 처리 후" 1회만 (검증 실패/오류 시 소스 유지 = 재시도 가능)
"""
import re, glob, shutil, sys, os, csv
from pathlib import Path
from datetime import date
from openpyxl import load_workbook

# XML 1.0에서 허용되지 않는 제어문자 제거 (openpyxl IllegalCharacterError 방지)
_ILLEGAL_CHARS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f'
    r'\ud800-\udfff\ufdd0-\ufddf\ufffe\uffff]'
)

def sanitize(val):
    if isinstance(val, str):
        return _ILLEGAL_CHARS_RE.sub('', val)
    return val

# ===================== CONFIG (로컬 경로) =====================
# 소스 폴더 = 스카이님 공유용 스프레드 폴더 (CSV가 여기로 떨어짐)
SOURCE_DIR    = r"G:\.shortcut-targets-by-id\1aExMnOUaz0KyUTRAhiSvAjCebgHx7Wa1\스카이님 공유용 스프레드 개설"
SOURCE_GLOBS  = ["유튜브_인플루언서_리스트*.xlsx", "유튜브_인플루언서_리스트*.csv"]

# 타겟 1: 마스터 DB (스케줄 폴더, "0." prefix 우선)
TARGET_DIR    = r"C:\Users\user\비서\스케줄"
TARGET_GLOB   = "*유튜브*인플루언서*관리*.xlsx"
# 타겟 2: 공유 파일 (소스 폴더 안에 같이 있음)
SHARED_DIR    = SOURCE_DIR
SHARED_GLOB   = "유튜브 인플루언서 관리_공유_*.xlsx"

# 처리할 타겟 목록 (라벨, 폴더, glob, 제외 키워드)
TARGETS = [
    {"label": "마스터", "dir": TARGET_DIR, "glob": TARGET_GLOB, "exclude": ("_backup_", "백업")},
    {"label": "공유",   "dir": SHARED_DIR, "glob": SHARED_GLOB, "exclude": ("_backup_", "백업")},
]

BACKUP_DIR    = r"C:\Users\user\비서\스케줄\old 및 백업"   # 타겟 백업 저장 위치
OLD_DIR       = r"C:\Users\user\Desktop\■■ 인플채널 wBS ■■\유튜브 DB 수집 프로그램 제작\인플루언서 리스트 old"
SHEET_NAME    = "메일발송현황"
NORMALIZE_COUNTS = False   # True 로 바꾸면 구독자수/조회수/영상수의 "1.2만","1,234" 등을 숫자로 변환
# ============================================================

ALIASES = {
    "키워드":   ["키워드", "검색어", "keyword"],
    "채널명":   ["채널명", "채널", "channel", "channelname", "채널이름"],
    "링크":     ["링크", "채널링크", "채널url", "채널주소", "url", "link"],
    "구독자수": ["구독자수", "구독자", "subscribers", "subscriber"],
    "조회수":   ["조회수", "총조회수", "views", "viewcount"],
    "영상수":   ["영상수", "동영상수", "videos", "videocount"],
    "연락처":   ["연락처", "이메일", "email", "메일", "이메일주소", "contact"],
    "블로그":   ["블로그", "blog", "네이버블로그", "블로그링크", "블로그url"],
    "카페":     ["카페", "cafe", "네이버카페", "카페링크", "카페url"],
    "인스타":   ["인스타", "인스타그램", "instagram", "insta", "ig", "인스타링크"],
    "쓰레드":   ["쓰레드", "스레드", "threads", "thread", "쓰레드링크"],
}
COUNT_FIELDS = {"구독자수", "조회수", "영상수"}


def norm(s):
    if s is None:
        return ""
    return re.sub(r"[\s_\-()\[\].·]", "", str(s)).lower()

# norm(variant) -> canonical
ALIAS_LOOKUP = {}
for canon, variants in ALIASES.items():
    for v in variants:
        ALIAS_LOOKUP[norm(v)] = canon


def pick_sources(source_dir, explicit_path=None):
    """모든 소스 파일 반환. explicit_path 지정 시 해당 파일만."""
    if explicit_path:
        if not Path(explicit_path).exists():
            sys.exit(f"[중단] 지정 소스 파일 없음: {explicit_path}")
        return [explicit_path]
    candidates = []
    for pattern in SOURCE_GLOBS:
        found = [f for f in glob.glob(str(Path(source_dir) / pattern))
                 if not Path(f).name.startswith("~$")]
        candidates.extend(found)
    if not candidates:
        sys.exit(f"[중단] 소스 파일 없음 ({source_dir} 에 유튜브_인플루언서_리스트*.csv/xlsx 필요)")
    candidates.sort(key=lambda f: Path(f).stat().st_mtime)
    print(f"[소스] {len(candidates)}개 파일 처리 예정: " + ", ".join(Path(f).name for f in candidates))
    return candidates


def _target_sort_key(path):
    """타겟 선택 우선순위: 1) '0.' prefix 여부  2) 파일명 내 YYMMDD 숫자  3) r뒤 숫자"""
    name = Path(path).name
    has_zero = name.startswith("0.")
    date_m = re.search(r'(\d{6})', name)
    date_val = int(date_m.group(1)) if date_m else 0
    r_m = re.search(r'[_\s]r(\d+)', name, re.IGNORECASE)
    r_val = int(r_m.group(1)) if r_m else 0
    return (has_zero, date_val, r_val)


def pick_target(target_dir, target_glob, exclude_substr=()):
    files = [f for f in glob.glob(str(Path(target_dir) / target_glob))
             if not Path(f).name.startswith("~$")
             and not any(x in Path(f).name for x in exclude_substr)]
    if not files:
        sys.exit(f"[중단] 타겟 파일을 못 찾음: {target_dir}\\{target_glob}")
    files.sort(key=_target_sort_key, reverse=True)
    if len(files) > 1:
        print(f"[주의] 타겟 후보 {len(files)}개 -> 자동 선택: {Path(files[0]).name}")
    return files[0]


def detect_header_row(rows, scan=15):
    """채널명이 있고 다른 알려진 필드가 2개 이상인 행을 헤더로 판정."""
    for i, row in enumerate(rows[:scan]):
        canon = {ALIAS_LOOKUP.get(norm(c)) for c in row if c is not None}
        canon.discard(None)
        if "채널명" in canon and len(canon) >= 3:
            return i
    return None


def field_of(header_text):
    return ALIAS_LOOKUP.get(norm(header_text))


def normalize_count(v):
    if v is None or isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    m = re.match(r"^([\d.]+)\s*(억|만|천|k|m|b)?\s*$", s, re.I)
    if not m:
        return v
    num = float(m.group(1)); unit = (m.group(2) or "").lower()
    mult = {"": 1, "천": 1e3, "만": 1e4, "억": 1e8, "k": 1e3, "m": 1e6, "b": 1e9}
    val = num * mult[unit]
    return int(val) if val == int(val) else val


def _empty(v):
    return v is None or str(v).strip() == ""


def read_source_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        h = detect_header_row(rows)
        if h is None:
            continue
        headers = rows[h]
        col_field = {ci: field_of(hv) for ci, hv in enumerate(headers)}
        records = []
        for r in rows[h + 1:]:
            rec = {}
            for ci, fld in col_field.items():
                if fld and ci < len(r):
                    rec[fld] = r[ci]
            if any(not _empty(rec.get(k)) for k in ("채널명", "링크", "키워드")):
                records.append(rec)
        wb.close()
        return records, {f for f in col_field.values() if f}
    wb.close()
    sys.exit("[중단] 소스 xlsx에서 헤더(채널명/키워드 포함 행)를 못 찾음")


def read_source_csv(path):
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        try:
            with open(path, encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                col_field = {h: field_of(h) for h in headers}
                records = []
                for row in reader:
                    rec = {fld: (row[h] if row[h] else None)
                           for h, fld in col_field.items() if fld}
                    if any(not _empty(rec.get(k)) for k in ("채널명", "링크", "키워드")):
                        records.append(rec)
            return records, {f for f in col_field.values() if f}
        except UnicodeDecodeError:
            continue
    sys.exit("[중단] CSV 파일 인코딩 인식 불가 (utf-8-sig/utf-8/cp949/euc-kr 모두 실패)")


def read_source(path):
    if Path(path).suffix.lower() == ".csv":
        return read_source_csv(path)
    return read_source_xlsx(path)


def col_index_map(ws, header_row):
    """target sheet: canonical field -> column index(1-based)"""
    out = {}
    for cell in ws[header_row]:
        fld = field_of(cell.value)
        if fld and fld not in out:
            out[fld] = cell.column
    return out


def last_data_row(ws, header_row, key_col):
    last = header_row
    for row in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=row, column=key_col).value not in (None, ""):
            last = row
    return last


def existing_keys(ws, header_row, last_row, ch_col, link_col):
    keys = set()
    for row in range(header_row + 1, last_row + 1):
        ch = ws.cell(row=row, column=ch_col).value
        lk = ws.cell(row=row, column=link_col).value if link_col else None
        keys.add((norm(ch), norm(lk)))
    return keys


def append_source_to_target(source_path, target_path, cmap, header_row,
                            ch_col, link_col, exist, current_last):
    """소스 1개를 타겟에 append (이동/백업은 하지 않음).
    반환: (new_last, exist, added, ok)"""
    sp = Path(source_path)
    records, src_fields = read_source(source_path)

    only_src = src_fields - set(cmap)
    if only_src:
        print(f"    [주의] 소스에만 있고 타겟에 없는 필드(누락): {sorted(only_src)}")

    seen, to_add, dup_exist, dup_self = set(), [], 0, 0
    for rec in records:
        key = (norm(rec.get("채널명")), norm(rec.get("링크")))
        if key in exist:
            dup_exist += 1; continue
        if key in seen:
            dup_self += 1; continue
        seen.add(key); to_add.append(rec)

    if not to_add:
        print(f"    {sp.name}: 소스 {len(records)}행 -> 신규 0행 (기존중복 {dup_exist}, 내부중복 {dup_self})")
        return current_last, exist, 0, True

    wb = load_workbook(target_path)
    ws = wb[SHEET_NAME]
    start = current_last + 1
    before = current_last - header_row
    for i, rec in enumerate(to_add):
        r = start + i
        for fld, col in cmap.items():
            if fld in rec:
                val = rec[fld]
                if NORMALIZE_COUNTS and fld in COUNT_FIELDS:
                    val = normalize_count(val)
                ws.cell(row=r, column=col).value = sanitize(val)
    wb.save(target_path)

    wb2 = load_workbook(target_path)
    ws2 = wb2[SHEET_NAME]
    new_last = last_data_row(ws2, header_row, ch_col)
    after = new_last - header_row
    ok = (after == before + len(to_add))
    wb2.close()

    status = "OK" if ok else "검증불일치!"
    print(f"    {sp.name}: 소스 {len(records)}행 -> 신규 {len(to_add)}행 추가 (행 {start}~{new_last}) [{status}]")

    if not ok:
        return current_last, exist, 0, False

    for rec in to_add:
        exist.add((norm(rec.get("채널명")), norm(rec.get("링크"))))
    return new_last, exist, len(to_add), True


def backup_target(target_path, today):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    bkp = Path(BACKUP_DIR) / f"{Path(target_path).stem}_backup_{today}.xlsx"
    n = 1
    while bkp.exists():
        bkp = Path(BACKUP_DIR) / f"{Path(target_path).stem}_backup_{today}({n}).xlsx"; n += 1
    shutil.copy2(target_path, bkp)
    print(f"[백업] {bkp.name}")


def process_target(spec, sources, today):
    """타겟 1개 처리: 백업 -> 구조파악 -> 소스 전체 append. 반환 요약 dict."""
    label = spec["label"]
    target_path = pick_target(spec["dir"], spec["glob"], exclude_substr=spec["exclude"])
    print(f"\n===== [{label}] {Path(target_path).name} =====")

    backup_target(target_path, today)

    wb0 = load_workbook(target_path)
    if SHEET_NAME not in wb0.sheetnames:
        sheets = wb0.sheetnames; wb0.close()
        print(f"[건너뜀] '{SHEET_NAME}' 시트 없음. 보유 시트: {sheets}")
        return {"label": label, "ok": False, "added": 0, "before": 0, "after": 0}
    ws0 = wb0[SHEET_NAME]
    grid = [[c.value for c in row] for row in ws0.iter_rows(min_row=1, max_row=15)]
    hr = detect_header_row(grid)
    header_row = (hr + 1) if hr is not None else 6
    cmap = col_index_map(ws0, header_row)
    if "채널명" not in cmap:
        wb0.close()
        print("[건너뜀] 타겟에 '채널명' 컬럼을 못 찾음")
        return {"label": label, "ok": False, "added": 0, "before": 0, "after": 0}
    ch_col = cmap["채널명"]; link_col = cmap.get("링크")
    current_last = last_data_row(ws0, header_row, ch_col)
    exist = existing_keys(ws0, header_row, current_last, ch_col, link_col)
    wb0.close()
    before = current_last - header_row
    print(f"[타겟] 헤더행 {header_row}, 컬럼 " + ",".join(f"{f}:{c}" for f, c in sorted(cmap.items(), key=lambda x: x[1])))
    print(f"[타겟] 기존 데이터 {before}행")

    total = 0; all_ok = True
    try:
        for src in sources:
            current_last, exist, added, ok = append_source_to_target(
                src, target_path, cmap, header_row, ch_col, link_col, exist, current_last)
            total += added
            if not ok:
                all_ok = False
    except PermissionError:
        print(f"[오류] 타겟 파일 잠김(Excel에서 열려있는지 확인): {Path(target_path).name}")
        all_ok = False

    after = current_last - header_row
    print(f"[{label}] {before} -> {after} (신규 {total}행) " + ("OK" if all_ok else "※ 검증/오류 있음"))
    return {"label": label, "ok": all_ok, "added": total, "before": before, "after": after}


def move_sources(sources, today):
    os.makedirs(OLD_DIR, exist_ok=True)
    for src in sources:
        sp = Path(src)
        dest = Path(OLD_DIR) / f"{sp.stem}_{today}완료{sp.suffix}"
        n = 1
        while dest.exists():
            dest = Path(OLD_DIR) / f"{sp.stem}_{today}완료({n}){sp.suffix}"; n += 1
        shutil.move(src, dest)
        print(f"[이동] {sp.name} -> {dest.name}")


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=None, help="소스 파일 경로 직접 지정 (생략 시 소스 폴더 전체)")
    parser.add_argument("--no-move", action="store_true", help="처리 후 소스 파일 이동 안 함")
    args = parser.parse_args()

    today = date.today().strftime("%y%m%d")

    # 소스 목록 (전체) — 모든 타겟이 동일 소스를 공유
    sources = pick_sources(SOURCE_DIR, explicit_path=args.source)

    # 타겟별 처리 (소스 이동은 아직 안 함)
    results = []
    for spec in TARGETS:
        results.append(process_target(spec, sources, today))

    all_ok = all(r["ok"] for r in results)

    print("\n========== 요약 ==========")
    for r in results:
        print(f"  {r['label']}: {r['before']} -> {r['after']} (+{r['added']}행) {'OK' if r['ok'] else '문제있음'}")

    # 소스 이동: 모든 타겟이 정상일 때만 1회
    if all_ok and not args.no_move:
        print("\n[소스 이동] 전 타겟 정상 처리 -> 소스 보관 폴더로 이동")
        move_sources(sources, today)
    elif not all_ok:
        print("\n[소스 유지] 일부 타겟 검증/오류 -> 소스 이동 안 함 (백업으로 복구 가능, 재시도 가능)")
    else:
        print("\n[소스 유지] --no-move 옵션 지정됨")

    print("[완료]")


if __name__ == "__main__":
    main()
