#!/usr/bin/env python3
# parse_order.py — STEP 3-4: 주문 파일 복호화 + 버킷 분류 + Raw_Data 반영
# 사용법: python parse_order.py <암호화_xlsx_경로> <managed_set_json> [--import-mode]
# 출력: stdout JSON (버킷 분류 결과 + 신규 건수)
import sys
import json
import re
import io
import shutil
import tempfile
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import msoffcrypto
import openpyxl
from dotenv import load_dotenv
import os

BASE_DIR     = Path(r"C:\Users\user\비서")
INPUT_DIR    = BASE_DIR / "input"
OUTPUT_DIR   = BASE_DIR / "output"
RAWDATA_PATH = BASE_DIR / "스케줄" / "정산DB_업데이트.xlsx"
CONFIG_PATH  = BASE_DIR / ".claude" / "skills" / "settlement-generator" / "scripts" / "ytber_config.json"
SKIP_LOG     = OUTPUT_DIR / "settlement_skipped.log"

DECRYPTED_PATH = INPUT_DIR / "order_decrypted.xlsx"
ORDER_SHEET    = "주문조회"
RAWDATA_SHEET  = "Raw_Data"

# 상품명에서 유튜버명 추출 패턴
YTBER_PATTERN = re.compile(r"\[([^\]]+?)\s*구독자")

# 주문 파일 열 인덱스 (0-based) — 실제 스마트스토어 주문조회 파일 기준
COL_ORDER_NO     = 0   # 상품주문번호
COL_ORDER_NO2    = 1   # 주문번호
COL_ORDER_DATE   = 2   # 주문일시
COL_ORDER_STATUS = 3   # 주문상태
COL_DELIVERY     = 4   # 배송속성
COL_CLAIM_STATUS = 6   # 클레임상태
COL_CLAIM_QTY    = 7   # 수량클레임 여부
COL_PRODUCT_ID   = 8   # 상품번호
COL_PRODUCT_NAME = 9   # 상품명
COL_OPTION1      = 10  # 옵션정보
COL_OPTION2      = 11  # 판매옵션정보
COL_QTY          = 12  # 수량
COL_BUYER_NAME   = 13  # 구매자명
COL_BUYER_ID     = 14  # 구매자ID
COL_AMOUNT       = -1  # 주문금액 (이 파일에는 없음 → 0 처리)


def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def decrypt_xlsx(encrypted_path: Path, password: str) -> Path:
    with open(encrypted_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        decrypted = io.BytesIO()
        office_file.decrypt(decrypted)

    DECRYPTED_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(DECRYPTED_PATH, "wb") as out:
        out.write(decrypted.getvalue())

    print(f"[INFO] 복호화 완료: {DECRYPTED_PATH}", file=sys.stderr)
    return DECRYPTED_PATH


def get_existing_order_nos(rawdata_wb) -> set:
    if RAWDATA_SHEET not in rawdata_wb.sheetnames:
        return set()
    ws = rawdata_wb[RAWDATA_SHEET]
    nos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            nos.add(str(row[0]).strip())
    return nos


def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def safe_int(val, default=0) -> int:
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def extract_ytber(product_name: str, name_map: dict) -> str | None:
    m = YTBER_PATTERN.search(product_name)
    if not m:
        return None
    raw_name = m.group(1).strip()
    return name_map.get(raw_name, raw_name)


def classify_product_store(product_id: str, product_name: str, registry: dict,
                           file_store: str | None = None) -> tuple[str, str | None]:
    """주문 1건을 (제품키, 스토어키) 로 분류.
    제품: 상품번호(by_product_id) 우선 → 상품명 키워드(by_keyword) 보조 → 기본값.
    스토어 우선순위: 상품번호 매핑 > 파일 스토어(file_store, --store) > 기본값.
      흑염소는 양쪽 스토어 상품명이 같으므로 스토어는 키워드가 아니라
      '어느 스토어 주문 파일이냐(file_store)' 또는 상품번호로만 구분한다."""
    default_product = registry.get("default_product", "흑염소")
    default_store   = file_store if file_store else registry.get("default_store", None)

    by_id = registry.get("by_product_id", {})
    if product_id and str(product_id) in by_id:
        m = by_id[str(product_id)]
        # 상품번호에 store가 명시되면 최우선, 아니면 파일 스토어/기본값 사용
        return m.get("product", default_product), m.get("store", default_store)

    for kw in registry.get("by_keyword", []):
        token = kw.get("contains", "")
        if token and token in (product_name or ""):
            return kw.get("product", default_product), default_store

    return default_product, default_store


def is_cancelled(order_status: str, claim_status: str) -> bool:
    return "취소" in order_status or "취소완료" in claim_status


def main():
    if len(sys.argv) < 3:
        print("사용법: python parse_order.py <암호화_xlsx> <managed_set_json>", file=sys.stderr)
        sys.exit(1)

    encrypted_path = Path(sys.argv[1])
    managed_set = set(json.loads(sys.argv[2]))

    # --store A|B : 이 주문 파일이 어느 스토어(판매자 계정)의 것인지 지정
    file_store = None
    if "--store" in sys.argv:
        i = sys.argv.index("--store")
        if i + 1 < len(sys.argv):
            file_store = (sys.argv[i + 1].strip().upper() or None)
            print(f"[INFO] 파일 스토어 지정: {file_store}", file=sys.stderr)

    load_dotenv(BASE_DIR / ".env")
    password = os.environ.get("SMARTSTORE_XLSX_PASSWORD", "")
    if not password:
        print("[ERROR] .env에 SMARTSTORE_XLSX_PASSWORD가 없습니다.", file=sys.stderr)
        sys.exit(1)

    config = load_config()
    name_map  = config.get("name_map", {})
    excludes  = set(config.get("exclude", []))
    registry  = config.get("product_registry", {})

    # 복호화
    try:
        decrypted_path = decrypt_xlsx(encrypted_path, password)
    except Exception as e:
        print(f"[ERROR] 복호화 실패: {e}", file=sys.stderr)
        sys.exit(1)

    # 주문 파일 로드
    try:
        order_wb = openpyxl.load_workbook(decrypted_path, data_only=True)
    except Exception as e:
        print(f"[ERROR] 복호화 파일 로드 실패: {e}", file=sys.stderr)
        sys.exit(1)

    # 주문 시트 선택: '주문조회'(구형) 우선 → '발주발송관리'(신형) → 첫 시트
    sheet_name = None
    for cand in (ORDER_SHEET, "발주발송관리"):
        if cand in order_wb.sheetnames:
            sheet_name = cand
            break
    if sheet_name is None:
        sheet_name = order_wb.sheetnames[0]
    order_ws = order_wb[sheet_name]
    print(f"[INFO] 주문 시트: {sheet_name}", file=sys.stderr)

    # 헤더 이름 기반 컬럼 매핑 (주문조회/발주발송관리 양식 모두 지원)
    all_rows = list(order_ws.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(all_rows[:6]):
        if r and any(c is not None and str(c).strip() == "상품주문번호" for c in r):
            header_idx = i
            break
    if header_idx is None:
        print(f"[ERROR] 헤더('상품주문번호') 행을 찾을 수 없음. 시트: {sheet_name}", file=sys.stderr)
        sys.exit(1)
    header = all_rows[header_idx]
    colmap = {str(h).strip(): j for j, h in enumerate(header) if h is not None and str(h).strip()}

    def cell(r, *names, default=""):
        for nm in names:
            j = colmap.get(nm)
            if j is not None and j < len(r) and r[j] is not None:
                return r[j]
        return default

    # Raw_Data 기존 주문번호 로드
    rawdata_wb = None
    existing_nos = set()
    if RAWDATA_PATH.exists():
        try:
            rawdata_wb = openpyxl.load_workbook(RAWDATA_PATH)
            existing_nos = get_existing_order_nos(rawdata_wb)
            print(f"[INFO] 기존 Raw_Data: {len(existing_nos)}건", file=sys.stderr)
        except Exception as e:
            print(f"[WARN] Raw_Data 읽기 실패, 신규 파일 생성: {e}", file=sys.stderr)
            rawdata_wb = openpyxl.Workbook()

    if rawdata_wb is None:
        rawdata_wb = openpyxl.Workbook()

    # Raw_Data 시트 준비
    # NOTE: 신규 컬럼(제품·스토어)은 기존 15개(인덱스 0~14) 뒤에 append.
    #       generate_sheets.py가 컬럼 인덱스로 읽으므로 순서를 절대 바꾸지 않는다.
    if RAWDATA_SHEET not in rawdata_wb.sheetnames:
        ws_raw = rawdata_wb.create_sheet(RAWDATA_SHEET)
        ws_raw.append([
            "상품주문번호", "주문번호", "주문일시", "주문상태", "배송속성",
            "유튜버 이름", "클레임상태", "수량클레임 여부", "상품번호", "상품명",
            "옵션정보", "판매옵션정보", "수량", "구매자명", "구매자ID",
            "제품", "스토어", "주문금액"
        ])
    else:
        ws_raw = rawdata_wb[RAWDATA_SHEET]
        # 기존 시트에 제품/스토어/주문금액 헤더가 없으면 16·17·18열(1-based)에 보강
        if not (ws_raw.cell(1, 16).value):
            ws_raw.cell(1, 16, "제품")
        if not (ws_raw.cell(1, 17).value):
            ws_raw.cell(1, 17, "스토어")
        if not (ws_raw.cell(1, 18).value):
            ws_raw.cell(1, 18, "주문금액")

    # 주문 분류
    buckets = {
        "settlement": [],   # 정산 대상
        "general":    [],   # 기타/일반
        "excluded":   [],   # 정산 제외 (미등재)
        "skipped":    [],   # 완전 제외 or 취소
    }
    unregistered = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_count = 0

    for row in all_rows[header_idx + 1:]:
        order_no = safe_str(cell(row, "상품주문번호"))
        if not order_no:
            continue

        order_no2    = safe_str(cell(row, "주문번호"))
        order_date   = safe_str(cell(row, "주문일시"))
        order_status = safe_str(cell(row, "주문상태"))
        delivery     = safe_str(cell(row, "배송속성", "배송방법"))
        claim_status = safe_str(cell(row, "클레임상태"))
        claim_qty    = safe_str(cell(row, "수량클레임 여부"))
        product_id   = safe_str(cell(row, "상품번호"))
        product_name = safe_str(cell(row, "상품명"))
        option1      = safe_str(cell(row, "옵션정보"))
        option2      = safe_str(cell(row, "판매옵션정보"))
        qty          = safe_int(cell(row, "수량"), 1)
        buyer_name   = safe_str(cell(row, "구매자명"))
        buyer_id     = safe_str(cell(row, "구매자ID"))
        # 실제 결제금액 (없으면 상품가격) — 비흑염소 제품 매출 집계에 사용
        amount       = safe_float(cell(row, "최종 상품별 총 주문금액", "상품가격", default=0))

        cancelled = is_cancelled(order_status, claim_status)

        # 유튜버명 추출 (취소 건도 포함해서 분류)
        ytber = extract_ytber(product_name, name_map)

        # 제품·스토어 분류 (제품: 상품번호→키워드 / 스토어: 상품번호→파일스토어→기본값)
        product_key, store_key = classify_product_store(product_id, product_name, registry, file_store)

        # 완전 제외 ytber는 Raw_Data 자체 차단
        if ytber in excludes:
            buckets["skipped"].append({"order_no": order_no, "ytber": ytber, "reason": "완전제외"})
            continue

        # 중복 체크 (취소 건도 중복이면 스킵)
        if order_no in existing_nos:
            continue

        ytber_label = ytber or config.get("general_sales_label", "기타/일반")

        default_product = registry.get("default_product", "흑염소")
        if cancelled:
            # 취소 건: Raw_Data 기록 + cancelled 버킷
            bucket = "cancelled"
            buckets["skipped"].append({"order_no": order_no, "ytber": ytber_label, "reason": "취소"})
        elif product_key != default_product:
            # 비흑염소 제품(화장품·수면영양제·올리브 등): Raw_Data엔 기록하되
            # 흑염소 정산/매출 버킷에는 넣지 않는다 (제품별 집계는 by_product가 담당)
            bucket = "other_product"
        elif ytber is None:
            bucket = "general"
        elif ytber in managed_set:
            bucket = "settlement"
        else:
            bucket = "excluded"
            unregistered.append({"name": ytber, "order_no": order_no, "qty": qty})
            _write_skip_log(ytber, qty, order_no)

        order_record = {
            "order_no":     order_no,
            "order_no2":    order_no2,
            "order_date":   order_date,
            "ytber":        ytber_label,
            "product_name": product_name,
            "qty":          qty,
            "buyer_name":   buyer_name,
            "amount":       amount,
            "bucket":       bucket,
            "is_cancelled": cancelled,
            "product":      product_key,
            "store":        store_key,
        }

        if cancelled:
            # 취소 건은 settlement/general 집계 제외, cancelled_by_ytber에 기록
            buckets.setdefault("cancelled_by_ytber", {}).setdefault(ytber_label, []).append(order_record)
        else:
            buckets.setdefault(bucket, []).append(order_record)

        existing_nos.add(order_no)
        new_count += 1

        # Raw_Data 반영 (양식 포맷: 스마트스토어 원본 컬럼 + 유튜버 이름)
        ws_raw.append([
            order_no, order_no2, order_date, order_status, delivery,
            ytber_label,
            claim_status, claim_qty, product_id, product_name,
            option1, option2, qty, buyer_name, buyer_id,
            product_key, (store_key or ""), int(amount)
        ])

    order_wb.close()

    # Raw_Data 저장
    RAWDATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    rawdata_wb.save(RAWDATA_PATH)
    rawdata_wb.close()
    print(f"[INFO] Raw_Data 저장: {RAWDATA_PATH} (신규 {new_count}건)", file=sys.stderr)

    result = {
        "new_count":          new_count,
        "settlement":         buckets["settlement"],
        "general":            buckets["general"],
        "excluded":           buckets["excluded"],
        "other_product":      buckets.get("other_product", []),
        "cancelled_by_ytber": buckets.get("cancelled_by_ytber", {}),
        "skipped_count":      len(buckets["skipped"]),
        "unregistered":       unregistered,
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


def _write_skip_log(ytber: str, qty: int, order_no: str):
    SKIP_LOG.parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = (
        f"[{now}] 정산 제외 기록\n"
        f"유튜버명: {ytber} | 주문번호: {order_no} | 수량: {qty}개 | 사유: 인플루언서관리 시트 미등재\n\n"
    )
    with open(SKIP_LOG, "a", encoding="utf-8") as f:
        f.write(line)


if __name__ == "__main__":
    main()
