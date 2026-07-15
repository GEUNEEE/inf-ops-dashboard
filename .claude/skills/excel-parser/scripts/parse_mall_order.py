#!/usr/bin/env python3
# parse_mall_order.py — 자사몰(spoteasy) 주문 zip 파싱 + Raw_Data 반영
# 사용법: python parse_mall_order.py <zip_경로>
# 출력: stdout JSON (parse_order.py와 동일한 버킷 스키마 — other_product 버킷 사용)
#
# zip 안의 CSV 컬럼(카페24 발주 다운로드 양식):
#   쇼핑몰, 쇼핑몰번호, 주문번호, 품목별 주문번호, 배송메시지, 총 주문금액, 총 결제금액,
#   상품번호, 주문상품명, 주문상품명(옵션포함), 수량, 판매가, 수령인, ...
# 자사몰은 화장품 전용 → 제품='화장품', 채널='자사몰' 고정.
# 정산예정금액 컬럼이 없으므로 빈값으로 기록 → 수익 계산 시 주문금액 기준 폴백(네이버와 동일 공식).
import sys
import json
import csv
import io
import os
import zipfile
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import openpyxl
from dotenv import load_dotenv

BASE_DIR     = Path(r"C:\Users\user\비서")
RAWDATA_PATH = BASE_DIR / "스케줄" / "정산DB_업데이트.xlsx"
RAWDATA_SHEET = "Raw_Data"

MALL_PRODUCT = "화장품"
MALL_CHANNEL = "자사몰"
GENERAL_LABEL = "기타/일반"


def read_zip_csv_rows(zip_path: Path, password: str) -> list[dict]:
    """zip 안의 모든 CSV를 읽어 dict 행 리스트로 반환."""
    rows = []
    with zipfile.ZipFile(zip_path) as z:
        names = [n for n in z.namelist() if n.lower().endswith(".csv")]
        if not names:
            raise RuntimeError("zip 안에 CSV 파일이 없습니다")
        for name in names:
            data = z.read(name, pwd=password.encode())
            for enc in ("utf-8-sig", "cp949", "utf-8"):
                try:
                    text = data.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise RuntimeError(f"CSV 인코딩 인식 실패: {name}")
            rows.extend(csv.DictReader(io.StringIO(text)))
    return rows


def get_existing_order_nos(rawdata_wb) -> set:
    if RAWDATA_SHEET not in rawdata_wb.sheetnames:
        return set()
    ws = rawdata_wb[RAWDATA_SHEET]
    nos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            nos.add(str(row[0]).strip())
    return nos


def safe_str(val) -> str:
    return "" if val is None else str(val).strip()


def safe_int(val, default=0) -> int:
    try:
        return int(float(val)) if val not in (None, "") else default
    except (ValueError, TypeError):
        return default


def extract_option(name: str, name_with_option: str) -> str:
    """주문상품명(옵션포함)에서 상품명 부분을 제거해 옵션 텍스트만 반환.
    세트 크기 판별(parse_set_size)이 'N개' 패턴을 읽을 수 있게 한다."""
    if name and name_with_option.startswith(name):
        return name_with_option[len(name):].strip()
    return name_with_option


def main():
    if len(sys.argv) < 2:
        print("사용법: python parse_mall_order.py <zip_경로>", file=sys.stderr)
        sys.exit(1)

    zip_path = Path(sys.argv[1])
    if not zip_path.exists():
        print(f"[ERROR] 파일 없음: {zip_path}", file=sys.stderr)
        sys.exit(1)

    load_dotenv(BASE_DIR / ".env")
    password = os.environ.get("JASAMALL_ZIP_PASSWORD", "")
    if not password:
        print("[ERROR] .env에 JASAMALL_ZIP_PASSWORD 없음 — 에스컬레이션", file=sys.stderr)
        sys.exit(1)

    try:
        csv_rows = read_zip_csv_rows(zip_path, password)
    except RuntimeError as e:
        # 비밀번호 오류 포함 — 파이프라인 중단 (에스컬레이션)
        print(f"[ERROR] zip 해제 실패: {e}", file=sys.stderr)
        sys.exit(1)

    rawdata_wb = openpyxl.load_workbook(RAWDATA_PATH)
    ws_raw = rawdata_wb[RAWDATA_SHEET]
    existing_nos = get_existing_order_nos(rawdata_wb)

    other_product = []
    new_count = 0
    seen_order2_amount = set()  # 같은 주문번호 다중 행 시 주문금액 이중 집계 방지

    for row in csv_rows:
        order_no  = safe_str(row.get("품목별 주문번호"))
        order_no2 = safe_str(row.get("주문번호"))
        if not order_no:
            continue
        if order_no in existing_nos:
            continue

        order_date  = safe_str(row.get("발주일"))
        product_id  = safe_str(row.get("상품번호"))
        name        = safe_str(row.get("주문상품명"))
        name_option = safe_str(row.get("주문상품명(옵션포함)"))
        option      = extract_option(name, name_option)
        qty         = safe_int(row.get("수량"), 1)
        buyer       = safe_str(row.get("수령인"))
        amount      = safe_int(row.get("총 주문금액"))

        # 총 주문금액은 주문 단위 금액 — 같은 주문번호의 두 번째 행부터는 0 처리
        if order_no2 in seen_order2_amount:
            print(f"[WARN] 주문 {order_no2} 다중 품목 — 두 번째 행부터 주문금액 0 처리, 수동 확인 필요", file=sys.stderr)
            amount = 0
        seen_order2_amount.add(order_no2)

        order_record = {
            "order_no":     order_no,
            "order_no2":    order_no2,
            "order_date":   order_date,
            "ytber":        GENERAL_LABEL,
            "product_name": name,
            "qty":          qty,
            "buyer_name":   buyer,
            "amount":       amount,
            "bucket":       "other_product",
            "is_cancelled": False,
            "product":      MALL_PRODUCT,
            "store":        "",
        }
        other_product.append(order_record)
        existing_nos.add(order_no)
        new_count += 1

        # Raw_Data 반영 — parse_order.py의 20컬럼 스키마와 동일
        ws_raw.append([
            order_no, order_no2, order_date, "결제완료", "",
            GENERAL_LABEL,
            "", "N", product_id, name,
            option, "", qty, buyer, "",
            MALL_PRODUCT, "", amount, None, MALL_CHANNEL
        ])

    rawdata_wb.save(RAWDATA_PATH)
    rawdata_wb.close()
    print(f"[INFO] 자사몰 주문 반영: {zip_path.name} (신규 {new_count}건)", file=sys.stderr)

    result = {
        "new_count":          new_count,
        "settlement":         [],
        "general":            [],
        "excluded":           [],
        "other_product":      other_product,
        "cancelled_by_ytber": {},
        "skipped_count":      0,
        "unregistered":       [],
    }
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
