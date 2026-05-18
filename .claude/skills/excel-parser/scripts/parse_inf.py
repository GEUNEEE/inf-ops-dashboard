#!/usr/bin/env python3
# parse_inf.py — STEP 2: 인플루언서관리 시트 → managed_set + 상태 집계 반환
# 출력: stdout JSON
import sys
import json
import re
from pathlib import Path
from datetime import date, datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import openpyxl

EXCEL_DIR   = Path(r"C:\Users\user\비서\스케줄")
SHEET_NAME  = "인플루언서관리"
CONFIG_PATH = Path(r"C:\Users\user\비서\.claude\skills\settlement-generator\scripts\ytber_config.json")

# 0-based 행 인덱스 (전치형: 행=항목, 열=인플루언서)
ROW_STATUS     = 9   # Excel Row 10: 현재상태
ROW_NAME       = 10  # Excel Row 11: 유튜버명
ROW_EXP_ACCEPT = 23  # Excel Row 24: 체험 수락일 (1차)
EXP_ROWS       = [23, 31, 34, 37]  # 1차 수락일 / 2·3·4차 체험 날짜
AD_ROWS        = [30, 33, 36, 39]  # Excel Row 31/34/37/40: 1~4차 광고 날짜

SPONSOR_COST_PER_EXP = 40000

STATUS_CATEGORIES = {
    "미팅대기": "미팅_대기",
    "미팅진행": "미팅_진행",
    "1차체험진행": "체험진행_1차",
    "2차체험진행": "체험진행_2차",
    "3차체험진행": "체험진행_3차",
    "1차광고예정": "광고예정_1차",
    "2차광고예정": "광고예정_2차",
    "1차광고완료": "광고완료_1차",
    "기타": "기타",
}


def load_name_map() -> dict:
    try:
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f).get("name_map", {})
    except Exception:
        return {}


def find_latest_excel(excel_dir: Path) -> Path:
    pattern = re.compile(r"인플루언서 관리_(\d{6})")
    candidates = []
    for f in excel_dir.glob("*.xlsx"):
        if f.name.startswith("~$") or "백업" in f.name:
            continue
        m = pattern.search(f.name)
        if m:
            candidates.append((int(m.group(1)), f.name, f))
    if not candidates:
        raise FileNotFoundError(f"'{excel_dir}'에서 마스터 DB xlsx를 찾을 수 없습니다.")
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def is_date(val) -> bool:
    from datetime import date, datetime
    if val is None:
        return False
    if isinstance(val, (datetime, date)):
        return True
    if isinstance(val, str) and val.strip():
        return True
    return False


def main():
    try:
        excel_path = find_latest_excel(EXCEL_DIR)
    except FileNotFoundError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] 마스터DB: {excel_path.name}", file=sys.stderr)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[ERROR] 엑셀 읽기 실패: {e}", file=sys.stderr)
        sys.exit(1)

    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] 시트 '{SHEET_NAME}' 없음. 목록: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    n_rows = len(rows)

    if ROW_NAME >= n_rows:
        print(f"[ERROR] 시트 행 수({n_rows}) < ROW_NAME({ROW_NAME})", file=sys.stderr)
        sys.exit(1)

    # 인플루언서 컬럼 탐색 (Row 11 비어있지 않은 첫 열부터)
    name_row = rows[ROW_NAME]
    start_col = 1
    while start_col < len(name_row) and not safe_str(name_row[start_col]):
        start_col += 1

    managed_set = []
    status_counter = {v: 0 for v in STATUS_CATEGORIES.values()}
    status_counter["기타"] = 0
    ad_total = 0
    ad_by_month = {}    # "YYYY-MM" → 광고 이벤트 수 (1~4차 합산)
    exp_by_month = {}   # "YYYY-MM" → 체험수락 인원 수

    name_map = load_name_map()
    per_influencer = {}

    for col_idx in range(start_col, len(name_row)):
        name = safe_str(name_row[col_idx])
        if not name:
            break

        status_raw = ""
        if ROW_STATUS < n_rows and col_idx < len(rows[ROW_STATUS]):
            status_raw = safe_str(rows[ROW_STATUS][col_idx])

        normalized_status = re.sub(r"\s+", "", status_raw)
        managed_set.append(name)

        category = STATUS_CATEGORIES.get(normalized_status, "기타")
        status_counter[category] = status_counter.get(category, 0) + 1

        normalized_name = name_map.get(name, name)

        # 체험 횟수 및 체험 날짜별 월 집계
        exp_cnt = 0
        exp_months = []  # 이 인플루언서의 체험 발생 월 목록 (중복 허용)
        for exp_row in EXP_ROWS:
            if exp_row < n_rows and col_idx < len(rows[exp_row]):
                val = rows[exp_row][col_idx]
                if is_date(val):
                    exp_cnt += 1
                    try:
                        if isinstance(val, (date, datetime)):
                            dt = val
                        else:
                            dt = datetime.strptime(str(val).strip(), "%Y-%m-%d")
                        exp_months.append(dt.strftime("%Y-%m"))
                    except Exception:
                        exp_months.append(None)

        # 인플루언서 관리 등록 = 최소 1회 체험진행 완료 기준
        exp_cnt = max(1, exp_cnt)

        per_influencer[normalized_name] = {
            "status": category,
            "exp_count": exp_cnt,
            "sponsor_cost": exp_cnt * SPONSOR_COST_PER_EXP,
            "exp_months": exp_months,  # 체험 발생 월 목록 (수익 월별 차감용)
        }

        # 체험 수락일 (Row 24, 1차) → 월별 인원 집계
        if ROW_EXP_ACCEPT < n_rows and col_idx < len(rows[ROW_EXP_ACCEPT]):
            val = rows[ROW_EXP_ACCEPT][col_idx]
            if val is not None:
                try:
                    if isinstance(val, (date, datetime)):
                        dt = val
                    else:
                        dt = datetime.strptime(str(val).strip(), "%Y-%m-%d")
                    ym = dt.strftime("%Y-%m")
                    exp_by_month[ym] = exp_by_month.get(ym, 0) + 1
                except Exception:
                    pass

        # 광고 전환: 1~4차 광고 날짜 각각 카운트 (날짜 기준 월별 집계)
        for ad_row in AD_ROWS:
            if ad_row >= n_rows or col_idx >= len(rows[ad_row]):
                continue
            val = rows[ad_row][col_idx]
            if not is_date(val):
                continue
            ad_total += 1
            try:
                if isinstance(val, (date, datetime)):
                    dt = val
                else:
                    dt = datetime.strptime(str(val).strip(), "%Y-%m-%d")
                ym = dt.strftime("%Y-%m")
                ad_by_month[ym] = ad_by_month.get(ym, 0) + 1
            except Exception:
                pass

    wb.close()

    # 체험 전환 총계 = 체험수락일 기재 인원 (인플루언서관리 시트 기준)
    exp_total = sum(exp_by_month.values())

    meeting_total = (
        status_counter.get("미팅_대기", 0)
        + status_counter.get("미팅_진행", 0)
    )

    result = {
        "managed_set": managed_set,
        "managed_count": len(managed_set),
        "inf_status": status_counter,
        "meeting_total": meeting_total,
        "exp_total": exp_total,
        "exp_by_month": dict(sorted(exp_by_month.items())),
        "ad_total": ad_total,
        "ad_by_month": dict(sorted(ad_by_month.items())),
        "per_influencer": per_influencer,
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"[INFO] 인플루언서관리 파싱 완료 — {len(managed_set)}명", file=sys.stderr)


if __name__ == "__main__":
    main()
