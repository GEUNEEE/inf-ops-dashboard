#!/usr/bin/env python3
# parse_mail.py — STEP 1: 메일발송현황 시트 파싱 → 퍼널 KPI 반환
# 출력: stdout JSON
import sys
import json
import re
from pathlib import Path
from datetime import date, datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import openpyxl

# 2026-08부터: 메일발송현황·인플루언서관리 둘 다 스카이님 공유 스프레드시트가 소스
# (발송 로그·상태값 모두 이쪽이 최신 — parse_inf.py도 동일 경로 사용)
_SHARED_DIR = Path(r"G:\.shortcut-targets-by-id\1aExMnOUaz0KyUTRAhiSvAjCebgHx7Wa1\스카이님 공유용 스프레드 개설")
_SHARED_PATTERN = re.compile(r"유튜브 인플루언서 관리_공유_(\d{6})")

def _find_shared_xlsx() -> Path:
    candidates = [(int(m.group(1)), f) for f in _SHARED_DIR.glob("*.xlsx")
                  if not f.name.startswith("~$") and "백업" not in f.name and (m := _SHARED_PATTERN.search(f.name))]
    if not candidates:
        raise FileNotFoundError(f"공유 스프레드시트를 찾을 수 없습니다: {_SHARED_DIR}")
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]

MAIL_SOURCE_XLSX = _find_shared_xlsx()
SHEET_NAME = "메일발송현황"

# 0-based 열 인덱스 (실제 메일발송현황 시트 기준)
COL_SEND_DATE   = 25  # Z  발송일
COL_REPLY_DATE  = 27  # AB 회신일
COL_STATUS      = 28  # AC 진행 상태
COL_MTG_DATE    = 29  # AD 미팅일
COL_ACCEPT_DATE = 30  # AE 협찬 수락일
COL_AD_DATE     = 31  # AF 광고 수락일
DATA_START_ROW  = 7   # 1-based

ETC_KEYWORDS = {"검토", "진행불가", "우리측거절", "기타"}


def cell_val(ws, row_1based, col_0based):
    try:
        return ws.cell(row=row_1based, column=col_0based + 1).value
    except Exception:
        return None


def is_date(val) -> bool:
    if val is None:
        return False
    if isinstance(val, (datetime, date)):
        return True
    if isinstance(val, str) and val.strip():
        return True
    return False


def normalize_status(raw) -> str:
    if raw is None:
        return ""
    return re.sub(r"\s+", "", str(raw).strip())


def main():
    try:
        excel_path = _find_shared_xlsx()
    except FileNotFoundError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
    if not excel_path.exists():
        print(f"[ERROR] 메일발송현황 소스 파일을 찾을 수 없습니다: {excel_path}", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] 메일발송현황 소스(공유 스프레드시트): {excel_path.name}", file=sys.stderr)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[ERROR] 엑셀 읽기 실패: {e}", file=sys.stderr)
        sys.exit(1)

    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] 시트 '{SHEET_NAME}' 없음. 시트 목록: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]

    total_sent = 0
    etc_count  = 0
    replied    = 0
    meeting    = 0
    exp_total  = 0
    ad_total   = 0
    empty_streak = 0
    by_month: dict = {}   # "YYYY-MM" → {sent, replied, meeting, exp, ad}

    def _ym(val):
        """날짜값 → 'YYYY-MM' 문자열, 파싱 불가 시 None"""
        if val is None:
            return None
        if isinstance(val, (datetime, date)):
            return val.strftime("%Y-%m")
        # 동기화로 "2026-07-01  12:00:00 AM"처럼 시간/AM-PM이 붙은 텍스트가 들어와도
        # 앞의 YYYY-MM-DD만 추출해 월 귀속시킨다 (뒷부분 시간 표기는 무시).
        m = re.match(r"(\d{4})-(\d{2})-\d{2}", str(val).strip())
        return f"{m.group(1)}-{m.group(2)}" if m else None

    def _bm(ym_key):
        """월별 집계 버킷 get-or-create"""
        if ym_key not in by_month:
            by_month[ym_key] = {"sent": 0, "replied": 0, "meeting": 0, "exp": 0, "ad": 0}
        return by_month[ym_key]

    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True)):
        if all(v is None for v in row_vals):
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0

        def _get(col_0based):
            return row_vals[col_0based] if col_0based < len(row_vals) else None

        send_date   = _get(COL_SEND_DATE)
        reply_date  = _get(COL_REPLY_DATE)
        status_raw  = _get(COL_STATUS)
        mtg_date    = _get(COL_MTG_DATE)
        accept_date = _get(COL_ACCEPT_DATE)
        ad_date     = _get(COL_AD_DATE)

        # 월별 집계는 '각 이벤트가 일어난 날짜'의 월로 귀속 (달력 기준).
        #   발송→발송일 / 응답→회신일 / 미팅→미팅일 / 체험→협찬수락일 / 광고→광고수락일
        # (예: 6월 발송·7월 미팅 건은 미팅이 7월에 잡힘)

        # 응답: 발송 여부와 무관하게 회신일 있으면 카운트, 월별은 회신일 기준
        # 회신일이 '필요' 등 날짜 아닌 텍스트여도 응답으로 집계, 월 귀속은 발송일로 폴백
        if is_date(reply_date):
            replied += 1
            ym_r = _ym(reply_date) or _ym(send_date)
            if ym_r:
                _bm(ym_r)["replied"] += 1

        # 이하 지표는 발송일 있는 행만
        if not is_date(send_date):
            continue

        total_sent += 1
        status = normalize_status(status_raw)
        ym_s = _ym(send_date)
        if ym_s:
            _bm(ym_s)["sent"] += 1

        # 미팅: 미팅일 있으면 카운트 (기타 포함), 월별은 미팅일 기준
        if is_date(mtg_date):
            meeting += 1
            ym_m = _ym(mtg_date)
            if ym_m:
                _bm(ym_m)["meeting"] += 1

        # 기타(검토/진행불가/우리측거절)는 체험·광고 지표에서만 제외
        if any(k in status for k in ETC_KEYWORDS) or status == "기타":
            etc_count += 1
            continue

        # 체험전환 (협찬수락일 기준 월)
        if is_date(accept_date):
            exp_total += 1
            ym_a = _ym(accept_date)
            if ym_a:
                _bm(ym_a)["exp"] += 1

        # 광고수락 (광고수락일 기준 월)
        if is_date(ad_date):
            ad_total += 1
            ym_d = _ym(ad_date)
            if ym_d:
                _bm(ym_d)["ad"] += 1

    wb.close()

    # 월별 비율 계산
    by_month_rates = {}
    for ym, d in sorted(by_month.items()):
        s = d["sent"] or 1
        by_month_rates[ym] = {
            "sent":         d["sent"],
            "replied":      d["replied"],
            "meeting":      d["meeting"],
            "exp":          d["exp"],
            "ad":           d["ad"],
            "reply_rate":   round(d["replied"] / s, 4),
            "meeting_rate": round(d["meeting"] / s, 4),
            "exp_rate":     round(d["exp"] / s, 4),
            "ad_rate":      round(d["ad"] / s, 4),
        }

    result = {
        "total_sent":       total_sent,
        "etc_excluded":     etc_count,
        "replied":          replied,
        "reply_rate":       round(replied / total_sent, 4) if total_sent else 0,
        "meeting_total":    meeting,
        "meeting_rate":     round(meeting / total_sent, 4) if total_sent else 0,
        "exp_total_approx": exp_total,
        "ad_total":         ad_total,
        "by_month":         by_month_rates,
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"[INFO] 메일발송현황 파싱 완료 — 총발송 {total_sent}건(기타포함), 기타 {etc_count}건", file=sys.stderr)


if __name__ == "__main__":
    main()
