#!/usr/bin/env python3
# generate_sheets.py — STEP 5: 정산서 생성 (송부용 - 템플릿 복사 방식)
# 사용법: python generate_sheets.py <bucket_json_path> <YYYY-MM>
# 출력: 유튜버별 월정산시트 작성 YYYY-MM_송부용.xlsx 생성 + stdout JSON (정산 요약)
import sys
import json
import shutil
import re
import calendar
from datetime import date
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.styles import GradientFill

BASE_DIR      = Path(r"C:\Users\user\비서")
RAWDATA_PATH  = BASE_DIR / "스케줄" / "정산DB_업데이트.xlsx"
CONFIG_PATH   = BASE_DIR / ".claude" / "skills" / "settlement-generator" / "scripts" / "ytber_config.json"
TEMPLATE_PATH = BASE_DIR / "스케줄" / "유튜버별 월정산시트 작성 4월분_송부용.xlsx"
LOGO_PATH     = BASE_DIR / ".claude" / "skills" / "settlement-generator" / "scripts" / "chobangli_logo.png"

# 정산DB Raw_Data 컬럼 인덱스 (0-based)
RAW_COL_DATE   = 2   # 주문일시
RAW_COL_STATUS = 3   # 주문상태
RAW_COL_YTBER  = 5   # 유튜버 이름
RAW_COL_CLAIM  = 6   # 클레임상태
RAW_COL_QTY    = 12  # 수량

# 템플릿 시트명 → ytber명 매핑 (처리 순서 포함)
TEMPLATE_SHEET_MAP = [
    ("마성민 4월",       "마성민"),
    ("나이스부부 4월",   "나이스부부"),
    ("C맹씨 4월",        "C맹씨"),
    ("The조치패밀리 4월", "The 조치 패밀리"),
    ("기타일반 4월",     "기타/일반"),
]

# 삭제할 구형 시트 (이전 월 히스토리 등)
OLD_SHEETS = ["Raw_Data3월", "마성민 3월", "나이스부부 3월", "유튜버별 정산 시트 만들기"]


def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def get_tier_price(cumulative_qty: int, tier_pricing: list) -> int:
    price = tier_pricing[0]["unit_price"]
    for tier in tier_pricing:
        if cumulative_qty >= tier["min_qty"]:
            price = tier["unit_price"]
    return price


def calc_monthly_amount(cum_before: int, ok_list: list, tier_pricing: list) -> tuple[int, int]:
    """단위 수량별 구간 적용: 구간 전 수량은 이전 단가, 경계 도달 시점부터 새 단가.
    Returns (total_settlement_amount, final_unit_price).
    """
    cum = cum_before
    total = 0
    final_price = get_tier_price(max(cum_before, 1), tier_pricing)
    for order in ok_list:
        for _ in range(order["qty"]):
            cum += 1
            price = get_tier_price(cum, tier_pricing)
            total += price
            final_price = price
    return total, final_price


def get_month_orders(rawdata_ws, target_month: str, name_map: dict = None) -> list:
    """정산DB Raw_Data에서 해당 월 주문 전체(취소 포함) 추출. name_map으로 ytber명 정규화."""
    nm = name_map or {}
    orders = []
    for row in rawdata_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        order_date_str = str(row[RAW_COL_DATE] or "")
        if not order_date_str.startswith(target_month):
            continue
        status = str(row[RAW_COL_STATUS] or "")
        claim  = str(row[RAW_COL_CLAIM]  or "")
        ytber_raw = str(row[RAW_COL_YTBER] or "")
        ytber = nm.get(ytber_raw, ytber_raw)
        orders.append({
            "order_no":     str(row[0]  or ""),
            "order_no2":    str(row[1]  or ""),
            "order_date":   order_date_str,
            "order_status": status,
            "delivery":     str(row[4]  or ""),
            "ytber":        ytber,
            "claim_status": claim,
            "claim_qty":    str(row[7]  or ""),
            "product_id":   str(row[8]  or ""),
            "product_name": str(row[9]  or ""),
            "option1":      str(row[10] or ""),
            "option2":      str(row[11] or ""),
            "qty":          int(row[RAW_COL_QTY]) if row[RAW_COL_QTY] else 0,
            "buyer_name":   str(row[13] or ""),
            "buyer_id":     str(row[14] or ""),
            "is_cancelled": ("취소" in status or "취소완료" in claim),
        })
    return orders


def get_cumulative_qty_before(rawdata_ws, ytber_name: str, target_month: str, name_map: dict = None) -> int:
    """해당 월 이전까지의 해당 ytber 비취소 누적 수량.
    Raw_Data에 없는 이전 월은 history/*.json에서 보완하여 정확한 누적을 보장."""
    nm = name_map or {}

    # ① Raw_Data에서 이전 월 수량 수집 (월별로 구분)
    raw_months_seen: set[str] = set()
    raw_total = 0
    for row in rawdata_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[RAW_COL_YTBER] is None:
            continue
        row_month = str(row[RAW_COL_DATE] or "")[:7]
        if row_month >= target_month:
            continue
        raw_months_seen.add(row_month)
        ytber_raw = str(row[RAW_COL_YTBER]).strip()
        if nm.get(ytber_raw, ytber_raw) != ytber_name:
            continue
        status = str(row[RAW_COL_STATUS] or "")
        claim  = str(row[RAW_COL_CLAIM]  or "")
        if "취소" in status or "취소완료" in claim:
            continue
        try:
            raw_total += int(row[RAW_COL_QTY]) if row[RAW_COL_QTY] else 0
        except (ValueError, TypeError):
            pass

    # ② Raw_Data에 없는 이전 월은 history 파일에서 보완
    HIST_DIR = Path(r"C:\Users\user\비서\site\data\history")
    hist_total = 0
    if HIST_DIR.exists():
        for hist_file in sorted(HIST_DIR.glob("*.json")):
            month = hist_file.stem  # "YYYY-MM"
            if month >= target_month:
                continue
            if month in raw_months_seen:
                continue  # Raw_Data에 이미 반영된 월은 스킵
            try:
                h = json.loads(hist_file.read_text(encoding="utf-8"))
                inf_data = h.get("influencers", {}).get(ytber_name, {})
                if not inf_data.get("is_general"):
                    hist_total += inf_data.get("qty", 0)
            except Exception:
                pass

    return raw_total + hist_total


_THIN       = Side(border_style="thin")
_MEDIUM     = Side(border_style="medium")
_NONE       = Side(border_style=None)
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")


def add_logo(ws):
    """초방리 마을 로고를 시트 B17 위치에 삽입"""
    if not LOGO_PATH.exists():
        return
    img = XLImage(str(LOGO_PATH))
    img.anchor = "B17"
    img.width  = 326
    img.height = 211
    ws.add_image(img)


def apply_table_style(ws, header_row: int, data_count: int, col_count: int):
    """주문 데이터 행에 테두리·흰색 fill·정렬 서식 적용 (헤더 포함)"""
    start_col = 2          # B
    end_col   = start_col + col_count - 1

    # 헤더행
    for col in range(start_col, end_col + 1):
        cell = ws.cell(header_row, col)
        cell.fill      = _WHITE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font      = Font(bold=True)
        left  = _MEDIUM if col == start_col else _THIN
        right = _MEDIUM if col == end_col   else _THIN
        cell.border = Border(top=_MEDIUM, bottom=_THIN, left=left, right=right)

    # 데이터행
    for i in range(data_count):
        row_num = header_row + 1 + i
        is_last = (i == data_count - 1)
        bottom  = _MEDIUM if is_last else _THIN

        for col in range(start_col, end_col + 1):
            cell  = ws.cell(row_num, col)
            left  = _MEDIUM if col == start_col else _THIN
            right = _MEDIUM if col == end_col   else _THIN
            cell.fill      = _WHITE_FILL
            cell.border    = Border(bottom=bottom, left=left, right=right)
            cell.alignment = Alignment(vertical="center")


def _sync_acct_from_excel(output_path: Path, ytber_info_map: dict, config: dict, config_path: Path):
    """기존 송부용 파일의 각 시트에서 계좌 정보를 읽어
    ytber_info_map에 없는 인플루언서를 채우고 ytber_config.json에 동기화한다.
    C4=파트너명, C6=판매링크, C9=계좌(이름\\n은행 계좌번호) 형식 파싱.
    """
    if not output_path.exists():
        return

    try:
        wb = openpyxl.load_workbook(str(output_path), data_only=True, read_only=True)
    except Exception as e:
        print(f"[WARN] 기존 정산 파일 열기 실패: {e}", file=sys.stderr)
        return

    newly_added = {}
    for sheet_name in wb.sheetnames:
        # 정산 시트만 처리 (월 시트명 패턴: "파트너명 N월")
        if not re.search(r"\d+월$", sheet_name):
            continue

        ws = wb[sheet_name]
        partner = ws["C4"].value
        if not partner or not str(partner).strip():
            continue
        partner = str(partner).strip()
        if partner in ytber_info_map:
            continue  # 이미 등록됨

        acct_raw = ws["C9"].value
        link     = ws["C6"].value

        # 계좌 정보가 없으면 스킵 (링크만 있어도 의미 없음)
        if not acct_raw or not str(acct_raw).strip():
            continue

        lines = [l.strip() for l in str(acct_raw).strip().splitlines() if l.strip()]
        if len(lines) < 2:
            continue  # "이름\n은행 계좌번호" 형식 아니면 스킵

        entry: dict = {
            "account_name": lines[0],
        }
        parts = lines[1].split(" ", 1)
        entry["account_bank"] = parts[0]
        entry["account_no"]   = parts[1] if len(parts) > 1 else ""
        if link:
            entry["smartstore_link"] = str(link).strip()

        ytber_info_map[partner] = entry
        newly_added[partner]    = entry
        print(f"[INFO] 계좌 정보 읽음 ({sheet_name}): {partner}", file=sys.stderr)

    wb.close()

    # ytber_config.json 동기화
    if newly_added:
        cfg_ytber = config.setdefault("ytber_info", {})
        cfg_ytber.update(newly_added)
        config_path.write_text(
            json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"[INFO] ytber_config.json 동기화: {list(newly_added.keys())}", file=sys.stderr)


def find_label_row(ws, label: str, search_col: int = 2) -> int | None:
    """search_col 열(1-based)에서 label이 있는 첫 번째 행 번호 반환"""
    for row in ws.iter_rows(min_col=search_col, max_col=search_col, values_only=False):
        if row[0].value == label:
            return row[0].row
    return None


def write_order_rows(ws, ok_list: list, can_list: list, is_general: bool) -> int:
    """'주문번호' 헤더 이하 기존 행을 삭제하고 ok/can 주문을 재기입 + 서식 적용.
    Returns: 마지막 데이터 행 번호 (인쇄 영역 계산용)"""
    header_row = find_label_row(ws, "주문번호")
    if header_row is None:
        return ws.max_row

    data_rows = ws.max_row - header_row
    if data_rows > 0:
        ws.delete_rows(header_row + 1, data_rows)

    for o in ok_list:
        if is_general:
            ws.append([None, o["order_no2"], o["product_name"], o["order_date"], o["buyer_name"], o["qty"], None])
        else:
            ws.append([None, o["order_no2"], o["product_name"], o["order_date"], o["qty"], None])

    for o in can_list:
        if is_general:
            ws.append([None, o["order_no2"], o["product_name"], o["order_date"], o["buyer_name"], o["qty"], "취소"])
        else:
            ws.append([None, o["order_no2"], o["product_name"], o["order_date"], o["qty"], "취소"])

    total = len(ok_list) + len(can_list)
    col_count = 6 if is_general else 5
    if total > 0:
        apply_table_style(ws, header_row, total, col_count)

    return header_row + total if total > 0 else header_row


def write_agg_values(ws, ok_list: list, can_list: list, settlement_amount, is_general: bool):
    """집계 값(건수·수량·금액)을 레이블 행의 C열에 직접 기입 (수식 대체)"""
    ok_qty = sum(o["qty"] for o in ok_list)

    label_map = {
        "총 주문 건수":          len(ok_list) + len(can_list),
        "주문 취소 건수":         len(can_list),
        "최종 판매 수량":         ok_qty,
        "최종 정산 금액 (세전)":  settlement_amount if (settlement_amount is not None and not is_general) else None,
        "최종 정산 금액":         settlement_amount if (settlement_amount is not None and not is_general) else None,
    }

    for label, value in label_map.items():
        if value is None:
            continue
        r = find_label_row(ws, label)
        if r:
            ws.cell(r, 3).value = value  # C열


def main():
    if len(sys.argv) < 3:
        print("사용법: python generate_sheets.py <bucket_json_path> <YYYY-MM>", file=sys.stderr)
        sys.exit(1)

    bucket_path      = sys.argv[1]
    settlement_month = sys.argv[2]  # "YYYY-MM"

    if bucket_path == "-":
        bucket_data = json.load(sys.stdin)
    else:
        with open(bucket_path, encoding="utf-8-sig") as f:
            bucket_data = json.load(f)

    year, month   = int(settlement_month[:4]), int(settlement_month[5:7])
    config        = load_config()
    tier_pricing  = config["tier_pricing"]
    general_label = config.get("general_sales_label", "기타/일반")
    ytber_info_map = dict(config.get("ytber_info", {}))  # 복사본으로 수정 허용
    name_map      = config.get("name_map", {})

    # output_path 미리 계산 (기존 파일 읽기에 필요)
    output_name = f"유튜버별 월정산시트 작성 {settlement_month}_송부용.xlsx"
    output_path = BASE_DIR / "스케줄" / output_name

    # 기존 송부용 파일에서 계좌 정보 읽기 → ytber_info_map 보완 + ytber_config.json 동기화
    _sync_acct_from_excel(output_path, ytber_info_map, config, CONFIG_PATH)

    # 정산DB 로드 (현월 주문 추출 + 누적 수량 계산)
    if not RAWDATA_PATH.exists():
        print(f"[ERROR] 정산DB 없음: {RAWDATA_PATH}", file=sys.stderr)
        sys.exit(1)

    rawdata_wb = openpyxl.load_workbook(RAWDATA_PATH, data_only=True, read_only=True)
    if "Raw_Data" not in rawdata_wb.sheetnames:
        print("[ERROR] 정산DB에 Raw_Data 시트 없음", file=sys.stderr)
        rawdata_wb.close()
        sys.exit(1)
    rawdata_ws = rawdata_wb["Raw_Data"]

    month_orders = get_month_orders(rawdata_ws, settlement_month, name_map)
    print(f"[INFO] {settlement_month} 주문 {len(month_orders)}건 (취소 포함)", file=sys.stderr)

    # ytber별 정상/취소 분류
    ytber_ok  = {}   # ytber → 비취소 주문 리스트
    ytber_can = {}   # ytber → 취소 주문 리스트
    for o in month_orders:
        ytber = o["ytber"]
        if o["is_cancelled"]:
            ytber_can.setdefault(ytber, []).append(o)
        else:
            ytber_ok.setdefault(ytber, []).append(o)

    # 해당 월 Raw_Data 전체 기반으로 동적 구성:
    # - 기타/일반 제외한 모든 비취소 ytber (batch 크기와 무관)
    # - 마지막에 기타/일반 추가
    settlement_ytbers = list(dict.fromkeys(
        ytber for ytber in ytber_ok
        if ytber != general_label
    ))
    dynamic_map = [(y, y) for y in settlement_ytbers]
    dynamic_map.append(("기타일반 4월", general_label))  # 기타/일반은 템플릿 시트명 고정
    # ytber_info에 없는 인플루언서는 계좌정보 없이 처리 (warn만)
    for ytber_name in settlement_ytbers:
        if ytber_name not in ytber_info_map:
            print(f"[WARN] ytber_info 없음 (계좌정보 누락): {ytber_name}", file=sys.stderr)

    # 템플릿 복사 (output_path는 위에서 이미 정의)
    shutil.copy(TEMPLATE_PATH, output_path)
    print(f"[INFO] 템플릿 복사: {output_path.name}", file=sys.stderr)

    # 출력 파일 열기 (수식 보존)
    wb = openpyxl.load_workbook(output_path)

    # ── Raw_Data 갱신 ──────────────────────────────────────────────────────
    ws_raw = wb["Raw_Data"]
    if ws_raw.max_row > 1:
        ws_raw.delete_rows(2, ws_raw.max_row - 1)
    for o in month_orders:
        ws_raw.append([
            o["order_no"], o["order_no2"], o["order_date"],
            o["order_status"], o["delivery"],
            o["ytber"],          # F열: 직접 ytber명 (수식 대신 값 기입)
            o["claim_status"], o["claim_qty"],
            o["product_id"], o["product_name"],
            o["option1"], o["option2"],
            o["qty"], o["buyer_name"], o["buyer_id"],
        ])

    # ── 정산 시트별 처리 ───────────────────────────────────────────────────
    summaries = []

    for tmpl_sheet, ytber_name in dynamic_map:
        if tmpl_sheet not in wb.sheetnames:
            # 템플릿에 없는 인플루언서: 마성민 시트를 복제해서 사용
            base_sheet = "마성민 4월"
            if base_sheet not in wb.sheetnames:
                print(f"[WARN] 기준 템플릿 시트 없음: {base_sheet}, {tmpl_sheet} 건너뜀", file=sys.stderr)
                continue
            wb.copy_worksheet(wb[base_sheet]).title = tmpl_sheet
            print(f"[INFO] 신규 시트 생성 (복제): {tmpl_sheet}", file=sys.stderr)

        ws = wb[tmpl_sheet]
        is_general = (ytber_name == general_label)
        ytber_info = ytber_info_map.get(ytber_name, {})

        # 기준 월 업데이트
        last_day = calendar.monthrange(year, month)[1]
        today_str = date.today().strftime("%Y-%m-%d")
        period_str = f"{year}-{month:02d}-01 ~ {year}-{month:02d}-{last_day}"

        if is_general:
            ws["C17"] = month   # 기타일반 시트는 C17 참조
        else:
            ws["I4"] = month    # 일반 인플루언서 시트는 I4 참조
            ws["B2"] = f"{year}년 {month}월 광고수익 정산 내역서"
            ws["C4"] = ytber_name
            ws["C7"] = period_str
            ws["C8"] = today_str

        # 개인 판매 링크 (모든 시트)
        link = ytber_info.get("smartstore_link", "")
        if link:
            ws["C6"] = link

        # 계좌 정보 (비기타 인플루언서만) — 없으면 빈칸으로 초기화
        if not is_general:
            acct_name = ytber_info.get("account_name", "")
            acct_bank = ytber_info.get("account_bank", "")
            acct_no   = ytber_info.get("account_no", "")
            if acct_name and acct_bank and acct_no:
                ws["C9"] = f"{acct_name}\n{acct_bank} {acct_no}"
            else:
                ws["C9"] = ""

        # 정산 단가 업데이트 (기타일반 제외)
        cum_before = 0
        unit_price = None
        if not is_general:
            cum_before = get_cumulative_qty_before(rawdata_ws, ytber_name, settlement_month, name_map)
            ws["C14"] = get_tier_price(cum_before, tier_pricing)  # 월 시작 시점 단가 (참고용)

        # 시트명 변경 (템플릿 4월 → 현재 월)
        safe_name  = re.sub(r'[\\/*?:\[\]/]', '', ytber_name)
        ws.title   = f"{safe_name} {month}월"[:31]

        # 집계 (summaries용)
        ok_list   = ytber_ok.get(ytber_name, [])
        can_list  = ytber_can.get(ytber_name, [])
        final_qty = sum(o["qty"] for o in ok_list)

        # 주문별 구간 단가 누적 계산 (경계 주문부터 새 단가 적용)
        if not is_general:
            amount, unit_price = calc_monthly_amount(cum_before, ok_list, tier_pricing)
            ws["C14"] = unit_price  # 최종 도달 단가로 갱신
        else:
            amount, unit_price = None, None

        cum_total = (cum_before + final_qty) if not is_general else None

        # 주문 상세 행 + 집계 값 직접 기입 (수식 의존 제거)
        last_data_row = write_order_rows(ws, ok_list, can_list, is_general)
        write_agg_values(ws, ok_list, can_list, amount, is_general)

        # 초방리 마을 로고 삽입
        add_logo(ws)

        # 인쇄 영역 + 1페이지 맞춤 + 보조 열(H-L) 숨김
        ws.print_area = f"$A$1:$G${last_data_row}"
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize   = 9   # A4
        for col_letter in ["H", "I", "J", "K", "L"]:
            ws.column_dimensions[col_letter].hidden = True

        summaries.append({
            "ytber":             ytber_name,
            "order_count":       len(ok_list),
            "qty":               final_qty,
            "cumulative_qty":    cum_total,
            "unit_price":        unit_price,
            "settlement_amount": amount,
            "is_general":        is_general,
        })
        print(
            f"[INFO] 정산 시트: {ytber_name} → {ws.title} "
            f"(정상 {len(ok_list)}건/{final_qty}개, 취소 {len(can_list)}건)",
            file=sys.stderr,
        )

    # ── 구형 시트 삭제 ─────────────────────────────────────────────────────
    for old in OLD_SHEETS:
        if old in wb.sheetnames:
            del wb[old]
            print(f"[INFO] 구형 시트 삭제: {old}", file=sys.stderr)

    rawdata_wb.close()

    wb.calculation.calcMode = "auto"
    wb.save(output_path)
    wb.close()
    print(f"[INFO] 저장 완료: {output_path}", file=sys.stderr)

    result = {
        "settlement_month": settlement_month,
        "output_path":      str(output_path),
        "sheet_count":      len(summaries),
        "summaries":        summaries,
        "unregistered":     bucket_data.get("unregistered", []),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
