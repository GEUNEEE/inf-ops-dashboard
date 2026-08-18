# -*- coding: utf-8 -*-
"""
델타 적용 작성기. compute_pull / compute_push 결과를 대상 파일에 반영.
원칙:
 - 쓰기 값은 SOURCE 파일에서 '원본 타입 그대로' 복사(날짜 보존). 델타의 cv문자열은 비교 전용.
 - 대상 쓰기 전 백업, 임시파일 교체 저장(원자적).
 - 충돌(conflict)은 적용하지 않는다(리포트만).
 - 인플관리 신규열의 수식행(체험종료/중간관리일)은 새 열 기준 수식으로 재작성.
"""
import os, re, shutil, sys
import datetime as _dt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from copy import copy
import sync_core as C

MAIL_LABEL2COL = {lab: i for i, lab in C.MAIL_EMP_COLS.items()}   # label -> 0-based col


class IntegrityError(RuntimeError):
    """적용 결과가 델타 범위를 벗어남(의도밖 변경/소실/신규). 대상 원본은 무변경."""
    def __init__(self, problems):
        self.problems = problems
        super().__init__("무결성 위반 %d건: %s" % (len(problems), "; ".join(problems[:5])))

_DATE_FMT_TOKEN = re.compile(r"[ymd]", re.I)
_TIME_FMT_TOKEN = re.compile(r"h|ss|AM/PM", re.I)


def _is_dt(v):
    return isinstance(v, (_dt.datetime, _dt.date, _dt.time))


def _cell_date_fmt(cell):
    """셀이 '날짜로 표시되는' 셀이면 그 number_format 반환, 아니면 None.
    datetime 값 외에 날짜서식 걸린 숫자(엑셀 시리얼, 예: 46158)도 포함."""
    v, fmt = cell.value, cell.number_format
    if v in (None, "") or fmt in (None, "General"):
        return None
    if _is_dt(v):
        return fmt
    if isinstance(v, (int, float)) and not isinstance(v, bool) and _DATE_FMT_TOKEN.search(fmt):
        return fmt
    return None


def _datelike(cell):
    """소스 셀이 날짜로 취급돼야 하는가 (datetime 값 또는 날짜서식 걸린 시리얼 숫자)."""
    return _is_dt(cell.value) or _cell_date_fmt(cell) is not None


def _pick_fmt(fmts):
    """수집된 기존 날짜 서식들 중 대표 서식 선택: 날짜전용(시각 토큰 없는) 서식 우선,
    그중 다수결(대소문자 무시). 과거 동기화가 남긴 'yyyy-mm-dd h:mm:ss' 같은
    광폭 서식(##### 원인)이 소수로 섞여 있어도 파일의 주류 양식을 따르게 함."""
    if not fmts:
        return None
    dateonly = [f for f in fmts if not _TIME_FMT_TOKEN.search(f)]
    pool = dateonly or fmts
    from collections import Counter
    best = Counter(f.lower() for f in pool).most_common(1)[0][0]
    for f in pool:
        if f.lower() == best:
            return f


def _col_date_fmt(ws, col, upto_row, start_row=7, limit=400, sample=30):
    """대상 시트 같은 컬럼에서 upto_row부터 위로 훑어 날짜 셀 서식을 최대 sample개
    수집 → 대표 서식 반환(=그 파일에 이미 입력된 날짜 양식). 없으면 None."""
    lo = max(start_row, upto_row - limit)
    fmts = []
    for r in range(upto_row, lo - 1, -1):
        f = _cell_date_fmt(ws.cell(row=r, column=col))
        if f:
            fmts.append(f)
            if len(fmts) >= sample:
                break
    return _pick_fmt(fmts)


def _row_date_fmt(ws, row, skip_col=None, start_col=3):
    """(인플관리 전치형) 같은 행의 기존 열에서 날짜 셀 서식을 수집 → 대표 서식 반환.
    수식 셀(체험종료/중간관리일)은 서식에 y/m/d 토큰이 있으면 날짜 서식으로 간주."""
    fmts = []
    for c in range(start_col, ws.max_column + 1):
        if c == skip_col:
            continue
        cell = ws.cell(row=row, column=c)
        v = cell.value
        f = _cell_date_fmt(cell)
        if f:
            fmts.append(f)
        elif (isinstance(v, str) and v.startswith("=")
                and cell.number_format not in (None, "General")
                and _DATE_FMT_TOKEN.search(cell.number_format)):
            fmts.append(cell.number_format)
    return _pick_fmt(fmts)


def _copy(dcell, scell, value=None, date_fmt=None):
    """원본 셀의 값과 표시서식(number_format)을 대상에 복사.
    단, 날짜값(또는 날짜 수식)이고 date_fmt(대상 파일의 기존 날짜 양식)가 주어지면
    소스 서식 대신 그걸 적용 → 소스/대상 날짜 양식 불일치로 ##### 표기되는 것 방지.
    value 지정 시 값만 대체(수식행 등)."""
    v = scell.value if value is None else value
    dcell.value = v
    if date_fmt and (_is_dt(v)
                     or (isinstance(v, (int, float)) and not isinstance(v, bool))
                     or (isinstance(v, str) and v.startswith("="))):
        dcell.number_format = date_fmt
    else:
        dcell.number_format = scell.number_format


def _inf_last_row(ws, cap=300):
    """인플관리 시트에서 값이 있는 마지막 행(최소 46). 47행 이후에도 인플 열에 정렬된
    데이터(광고 링크·판매·정산·비고 등)가 있으므로 열 이동/복사는 반드시 여기까지 포함."""
    last = 46
    maxr = min(ws.max_row or 46, cap)
    maxc = ws.max_column or 1
    for r in range(47, maxr + 1):
        for c in range(1, maxc + 1):
            if ws.cell(r, c).value not in (None, ""):
                last = r
                break
    return last


def _insert_inf_at_front(itgt, isrc, scol):
    """인플관리 3열에 isrc[scol] 인플을 삽입. 기존 열(3..N)을 오른쪽으로 한 칸씩 밀고
    값·서식·열너비 보존, 수식은 새 열참조로 변환. (직원이 공유본 맨 앞에 추가하는 패턴과 정렬)
    이동/복사 범위는 1행~데이터 마지막 행 전체 — 46행까지만 밀면 하단블록이 어긋나며 섞임."""
    NAME_ROW = C.INF_ROW_NAME
    cols = [c for c in range(3, itgt.max_column + 1)
            if itgt.cell(NAME_ROW, c).value not in (None, "")]
    ROWS = range(1, max(_inf_last_row(itgt), _inf_last_row(isrc)) + 1)
    row_fmt = {r: _row_date_fmt(itgt, r) for r in ROWS}   # 시프트 전 대상측 날짜 양식 채집
    if cols:
        last = max(cols)
        for c in range(last, 2, -1):           # last..3 → 오른쪽으로 이동(덮어쓰기 방지)
            sL, dL = get_column_letter(c), get_column_letter(c + 1)
            for r in ROWS:
                sc = itgt.cell(r, c)
                val = sc.value
                if isinstance(val, str) and val.startswith("="):
                    val = Translator(val, origin="%s%d" % (sL, r)).translate_formula("%s%d" % (dL, r))
                dc = itgt.cell(r, c + 1)
                dc.value = val
                dc._style = copy(sc._style)
            w = itgt.column_dimensions[sL].width
            if w is not None:
                itgt.column_dimensions[dL].width = w
    cl = get_column_letter(3)
    for r in ROWS:
        sc = isrc.cell(r, scol)
        if r in C.INF_FORMULA_ROWS:
            _copy(itgt.cell(r, 3), sc, value=("=%s27+10" % cl) if r == 28 else ("=%s28+7" % cl),
                  date_fmt=row_fmt.get(r))
        else:
            _copy(itgt.cell(r, 3), sc, date_fmt=row_fmt.get(r) if _datelike(sc) else None)
    # 검증: 3열에 신규 인플 이름이 들어갔는지(실패 시 예외→저장 생략, 백업 보존)
    if itgt.cell(NAME_ROW, 3).value in (None, ""):
        raise RuntimeError("inf front-insert 실패: 3열 이름 비어있음")

def _inf_index(ws):
    """{key: col(1-based)}, {label: row}"""
    rows = {r: ws.cell(row=r, column=1).value for r in range(1, 46)}
    label2row = {C.cv(rows[r]): r for r in C.INF_LABEL_ROWS}
    keymap = {}
    for col in range(3, ws.max_column + 1):
        name = ws.cell(row=C.INF_ROW_NAME, column=col).value
        if name in (None, ""):
            continue
        phone = C.norm_phone(ws.cell(row=C.INF_ROW_PHONE, column=col).value)
        key = phone if phone else "NAME::" + "".join(str(name).split())
        keymap[key] = col
    return keymap, label2row

def _mail_keyrow(ws):
    """{key: row(1-based)}. 중복 채널은 load_mail과 동일하게 2번째부터 '키#2','키#3'…"""
    out, es, seen = {}, 0, {}
    for r in range(7, ws.max_row + 1):
        name = ws.cell(row=r, column=5).value
        url = ws.cell(row=r, column=6).value
        if name in (None, "") and url in (None, ""):
            es += 1
            if es >= 15:
                break
            continue
        es = 0
        key = C.norm_url(C.cv(url)) if C.cv(url) else "NAME::" + "".join(C.cv(name).split())
        n = seen.get(key, 0) + 1
        seen[key] = n
        if n > 1:
            key = "%s#%d" % (key, n)
        out[key] = r
    return out

def _last_mail_row(ws):
    last = 6
    for r in range(7, ws.max_row + 1):
        if ws.cell(row=r, column=5).value or ws.cell(row=r, column=6).value:
            last = r
    return last

# ===================== 무결성 검증 =====================
def _mail_grid_ro(path):
    """{key: [34개 cv값]} — load_mail과 동일한 occurrence 키. read_only 로드(원본 무접촉)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[C.MAIL_SHEET]
    out, es, seen = {}, 0, {}
    for r in ws.iter_rows(min_row=7, values_only=True):
        if all(v is None for v in r):
            es += 1
            if es >= 15:
                break
            continue
        es = 0
        name = C.cv(r[4]) if len(r) > 4 else ""
        url = C.cv(r[5]) if len(r) > 5 else ""
        if name == "" and url == "":
            continue
        key = C.norm_url(url) if url else "NAME::" + "".join(name.split())
        n = seen.get(key, 0) + 1
        seen[key] = n
        if n > 1:
            key = "%s#%d" % (key, n)
        out[key] = [C.cv(r[i]) if i < len(r) else "" for i in range(34)]
    wb.close()
    return out


def _inf_grid_ro(path, cap=300):
    """{key: {행번호: cv값}} + {label: 행번호}. 열 이동(front-insert)과 무관한 키 단위 뷰.
    47행 이후 하단블록(광고 링크·판매·정산·비고 등)도 포함해 전 행 검증."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[C.INF_SHEET]
    rows = list(ws.iter_rows(min_row=1, max_row=cap, values_only=True))
    wb.close()
    last = 46
    for i, r in enumerate(rows):
        if i >= 46 and any(v not in (None, "") for v in r):
            last = i + 1
    label2row = {}
    for r in C.INF_LABEL_ROWS:
        if r - 1 < len(rows) and rows[r-1]:
            label2row[C.cv(rows[r-1][0])] = r
    name_row = rows[C.INF_ROW_NAME-1] if len(rows) >= C.INF_ROW_NAME else ()
    out = {}
    for col in range(2, len(name_row)):
        name = C.cv(name_row[col])
        if not name:
            continue
        phone_row = rows[C.INF_ROW_PHONE-1] if len(rows) >= C.INF_ROW_PHONE else ()
        phone = C.norm_phone(phone_row[col]) if col < len(phone_row) else ""
        key = phone if phone else "NAME::" + "".join(name.split())
        cells = {}
        for r in range(1, last + 1):
            v = rows[r-1][col] if (r - 1 < len(rows) and col < len(rows[r-1])) else None
            cells[r] = C.cv(v)
        out[key] = cells
    return out, label2row


def _eq(a, b):
    """cv 문자열 동등 비교. openpyxl 재저장이 정수값 float 표현을 정규화하는 경우
    ('2512.0'->'2512')는 수치 동등이면 같은 값으로 취급."""
    if a == b:
        return True
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return False


def verify_integrity(before_path, after_path, delta):
    """적용 전 파일 vs 적용 결과(임시파일)를 키 단위 전수 대조.
    델타에 명시된 셀 외의 변경·행/열 소실·의도밖 신규를 위반 목록으로 반환(비면 통과).
    front-insert 열 이동과 무관하도록 키 기준 비교, 수식 셀(=로 시작)은 열참조 변환 때문에 제외."""
    problems = []

    if delta.get("mail_append") or delta.get("mail_status"):
        old = _mail_grid_ro(before_path)
        new = _mail_grid_ro(after_path)
        ok_new = {x["key"] for x in delta.get("mail_append", [])}
        ok_upd = {}
        for x in delta.get("mail_status", []):
            c0 = MAIL_LABEL2COL.get(x["label"])
            if c0 is not None:
                ok_upd.setdefault(x["key"], set()).add(c0)
        for k, oc in old.items():
            nc = new.get(k)
            if nc is None:
                problems.append("메일 행 소실: %s" % k)
                continue
            for c in range(34):
                if not _eq(oc[c], nc[c]) and c not in ok_upd.get(k, ()):
                    if oc[c].startswith("=") or nc[c].startswith("="):
                        continue
                    problems.append("메일 의도밖 변경: %s col%d '%s'->'%s'" % (k, c, oc[c][:30], nc[c][:30]))
        for k in new:
            if k not in old and k not in ok_new:
                problems.append("메일 의도밖 신규행: %s" % k)

    if delta.get("inf_update") or delta.get("inf_add"):
        old, _ = _inf_grid_ro(before_path)
        new, lab2row = _inf_grid_ro(after_path)
        ok_new = {x["key"] for x in delta.get("inf_add", [])}
        ok_upd = {}
        for x in delta.get("inf_update", []):
            r = lab2row.get(x["label"])
            if r is not None:
                ok_upd.setdefault(x["key"], set()).add(r)
        for k, oc in old.items():
            nc = new.get(k)
            if nc is None:
                problems.append("인플 열 소실: %s" % k)
                continue
            for r in range(1, max(len(oc), len(nc)) + 1):
                ov, nv = oc.get(r, ""), nc.get(r, "")
                if not _eq(ov, nv) and r not in ok_upd.get(k, ()):
                    if r in C.INF_FORMULA_ROWS or ov.startswith("=") or nv.startswith("="):
                        continue
                    problems.append("인플 의도밖 변경: %s 행%d '%s'->'%s'" % (k, r, ov[:30], nv[:30]))
        for k in new:
            if k not in old and k not in ok_new:
                problems.append("인플 의도밖 신규: %s" % k)

    return problems


def apply_delta(source_path, target_path, delta, backup_dir=None, inf_add_front=False):
    """delta(pull 또는 push)를 target에 적용. source에서 원본값 복사. 적용 요약 반환.
    inf_add_front=True면 신규 인플을 target 맨 앞(3열)에 삽입(pull용). False면 끝에 append."""
    # 백업
    backup_dir = backup_dir or os.path.join(os.path.dirname(target_path), "old")
    os.makedirs(backup_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(target_path))[0]
    bkp = os.path.join(backup_dir, base + "_sync백업.xlsx")
    shutil.copy2(target_path, bkp)

    src = openpyxl.load_workbook(source_path, data_only=True)
    tgt = openpyxl.load_workbook(target_path)   # 수식/서식 보존
    applied = {"mail_append": 0, "mail_status": 0, "inf_update": 0, "inf_add": 0}

    # ===== 메일발송현황 =====
    if delta.get("mail_append") or delta.get("mail_status"):
        ssrc = src[C.MAIL_SHEET]; stgt = tgt[C.MAIL_SHEET]
        src_row = _mail_keyrow(ssrc)
        # 대상 시트에 이미 입력된 컬럼별 날짜 양식 (lazy 캐시)
        _cf_cache = {}
        last0 = _last_mail_row(stgt)
        def col_fmt(col1):
            if col1 not in _cf_cache:
                _cf_cache[col1] = _col_date_fmt(stgt, col1, last0)
            return _cf_cache[col1]
        # 신규행 append
        nr = last0 + 1
        for x in delta.get("mail_append", []):
            sr = src_row.get(x["key"])
            if sr:
                # 원본 행에서 값 복사, 날짜는 대상측 기존 양식으로 표기 통일
                for c in range(34):
                    sc = ssrc.cell(row=sr, column=c+1)
                    if sc.value is not None:
                        _copy(stgt.cell(row=nr, column=c+1), sc,
                              date_fmt=col_fmt(c+1) if _datelike(sc) else None)
            else:
                full = x.get("full")
                if not full:
                    continue
                for c, val in enumerate(full):
                    if val is not None:
                        stgt.cell(row=nr, column=c+1).value = val
            nr += 1
            applied["mail_append"] += 1
        # 상태 갱신
        tgt_row = _mail_keyrow(stgt)
        for x in delta.get("mail_status", []):
            col0 = MAIL_LABEL2COL.get(x["label"])
            sr = src_row.get(x["key"]); tr = tgt_row.get(x["key"])
            if col0 is None or sr is None or tr is None:
                continue
            sc = ssrc.cell(row=sr, column=col0+1)
            _copy(stgt.cell(row=tr, column=col0+1), sc,
                  date_fmt=col_fmt(col0+1) if _datelike(sc) else None)
            applied["mail_status"] += 1

    # ===== 인플루언서관리 =====
    if delta.get("inf_update") or delta.get("inf_add"):
        isrc = src[C.INF_SHEET]; itgt = tgt[C.INF_SHEET]
        skey, slab2row = _inf_index(isrc)
        tkey, tlab2row = _inf_index(itgt)
        # 대상 시트 행별 날짜 양식 (lazy 캐시, 전치형이라 행 단위)
        # 갱신 대상 셀 자신의 기존 서식도 선례에 포함(그 셀이 행의 유일한 날짜일 수 있음)
        _rf_cache = {}
        def row_fmt(row):
            if row not in _rf_cache:
                _rf_cache[row] = _row_date_fmt(itgt, row)
            return _rf_cache[row]
        # 셀 갱신
        for x in delta.get("inf_update", []):
            scol = skey.get(x["key"]); tcol = tkey.get(x["key"])
            row = tlab2row.get(x["label"])
            if scol is None or tcol is None or row is None:
                continue
            sc = isrc.cell(row=row, column=scol)
            _copy(itgt.cell(row=row, column=tcol), sc,
                  date_fmt=row_fmt(row) if _datelike(sc) else None)
            applied["inf_update"] += 1
        # 동명 인플 가드: 키(연락처)가 달라도 같은 이름이 이미 있으면 삽입 스킵(중복 열 방지)
        # — 2026-07-14 키 불일치로 동일 인플 3명이 양쪽에 중복 생성된 사고 재발 방지
        existing_names = {"".join(str(itgt.cell(C.INF_ROW_NAME, c).value).split())
                          for c in range(3, itgt.max_column + 1)
                          if itgt.cell(C.INF_ROW_NAME, c).value not in (None, "")}
        # 신규 인물 = 새 열. front-insert면 공유본 순서 유지 위해 역순으로 삽입(맨 앞에 차례로).
        adds = list(delta.get("inf_add", []))
        for x in (reversed(adds) if inf_add_front else adds):
            scol = skey.get(x["key"])
            if scol is None:
                continue
            sname = "".join(str(isrc.cell(C.INF_ROW_NAME, scol).value or "").split())
            if sname and sname in existing_names:
                applied.setdefault("inf_add_dup_skip", []).append(sname)
                continue
            if sname:
                existing_names.add(sname)
            if inf_add_front:
                _insert_inf_at_front(itgt, isrc, scol)
                tkey = {k: (c + 1 if c >= 3 else c) for k, c in tkey.items()}
                tkey[x["key"]] = 3
            else:
                ncol = max(tkey.values()) + 1 if tkey else 3
                cl = get_column_letter(ncol)
                # 끝열 append(push)는 1~45행만 복사 — 로컬 하단블록(47행+)은 과거 46행 시프트
                # 누락으로 열이 어긋나 있어, 정렬 복구 전까지 하단 복사 시 타 인플 데이터가 섞임
                for r in range(1, 46):
                    sc = isrc.cell(row=r, column=scol)
                    if r in C.INF_FORMULA_ROWS:
                        _copy(itgt.cell(row=r, column=ncol), sc,
                              value=("=%s27+10" % cl) if r == 28 else ("=%s28+7" % cl),
                              date_fmt=row_fmt(r))
                    else:
                        _copy(itgt.cell(row=r, column=ncol), sc,
                              date_fmt=row_fmt(r) if _datelike(sc) else None)
                tkey[x["key"]] = ncol
            applied["inf_add"] += 1

    # 임시파일 저장 -> 무결성 검증 통과 시에만 원본 교체 (위반 시 원본 무변경)
    tmp = target_path + ".synctmp.xlsx"
    tgt.save(tmp)
    src.close(); tgt.close()
    problems = verify_integrity(target_path, tmp, delta)
    if problems:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise IntegrityError(problems)
    os.replace(tmp, target_path)
    return applied, bkp
