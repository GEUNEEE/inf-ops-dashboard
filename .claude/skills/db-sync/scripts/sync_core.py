# -*- coding: utf-8 -*-
"""
db-sync 코어: 키 정규화 / 시트 로더 / 스냅샷 입출력.
공유본 <-> 로컬 3자 비교 델타 병합의 공용 기반.
어떤 함수도 원본을 저장하지 않는다(로더는 읽기전용). 쓰기는 호출 측에서 백업 후 수행.
"""
import os, re, json, glob

# ---------- 경로 ----------
SCHEDULE_DIR = r"C:\Users\user\비서\스케줄"
SHARED_DIR   = os.path.join(SCHEDULE_DIR, "스카이님 공유용 스프레드 개설")
SHARED_FILE  = os.path.join(SHARED_DIR, "1. 유튜브 인플루언서 관리_공유_260619.xlsx")
SNAP_DIR     = r"C:\Users\user\비서\output\sync_snapshot"

def latest_local():
    """스케줄 폴더의 '0. ...관리_*.xlsx' 중 최신(파일명 기준)."""
    cands = glob.glob(os.path.join(SCHEDULE_DIR, "0. *유튜브*인플루언서*관리*.xlsx"))
    cands = [c for c in cands if "~$" not in c]
    if not cands:
        raise FileNotFoundError("로컬 마스터를 찾을 수 없음: " + SCHEDULE_DIR)
    def key(p):
        m = re.search(r"_(\d{6})", os.path.basename(p))
        r = re.search(r"_r(\d+)", os.path.basename(p))
        return (m.group(1) if m else "", int(r.group(1)) if r else 0)
    return max(cands, key=key)

# ---------- 키 정규화 ----------
def norm_phone(v):
    return re.sub(r"\D", "", str(v)) if v is not None else ""

def norm_url(v):
    if v is None:
        return ""
    s = str(v).strip()
    m = re.search(r"(UC[0-9A-Za-z_\-]{20,})", s)
    if m:
        return m.group(1).lower()
    m = re.search(r"@([0-9A-Za-z._\-가-힣]+)", s)
    if m:
        return "@" + m.group(1).lower()
    s = re.sub(r"^https?://", "", s, flags=re.I)
    s = re.sub(r"^www\.", "", s, flags=re.I)
    return s.split("?")[0].rstrip("/").lower()

def cv(v):
    """셀 비교/직렬화용 정규화 문자열."""
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if isinstance(v, str):
        return v.strip()
    return str(v)

# ---------- 시트 좌표 ----------
# 메일발송현황 (표형식, 헤더 6행, 데이터 7행~) 0-based 컬럼 인덱스
MAIL_SHEET = "메일발송현황"
MAIL_NAME, MAIL_URL = 4, 5
MAIL_OWNER_COLS = {0: "분류"}                                      # 형 열(공유본에서 안 받음)
MAIL_EMP_COLS   = {25: "발송일", 26: "확인일", 27: "회신일",
                   28: "진행 상태", 29: "미팅일",
                   30: "협찬 수락일", 31: "광고 수락일"}            # 직원이 공유본에서 채우는 응답열(발송일·확인일 포함)

# 인플루언서관리 (전치형: A열 라벨, C열~ 인물). Excel 행번호(1-based)
INF_SHEET = "인플루언서관리"
INF_ROW_NAME, INF_ROW_PHONE = 11, 17
INF_FORMULA_ROWS = {28, 29}            # 체험 종료(=D27+10)·중간 관리일(=D28+7): 파생값, 동기화 제외
INF_LABEL_ROWS = [r for r in range(10, 46) if r not in INF_FORMULA_ROWS]
# 주의: load_inf의 'col'은 0-based(튜플 인덱스). openpyxl 쓰기 시 col+1 사용.

# ---------- 로더(읽기전용) ----------
def load_mail(path):
    """반환: {urlkey: {'name','url','row', 'emp': {label:val}, 'owner': {label:val}}}.
    중복 urlkey는 첫 행 기준."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[MAIL_SHEET]
    out, order, es = {}, [], 0
    rownum = 6
    for r in ws.iter_rows(min_row=7, values_only=True):
        rownum += 1
        if all(v is None for v in r):
            es += 1
            if es >= 15:
                break
            continue
        es = 0
        name = cv(r[MAIL_NAME]) if len(r) > MAIL_NAME else ""
        url  = cv(r[MAIL_URL]) if len(r) > MAIL_URL else ""
        if name == "" and url == "":
            continue
        key = norm_url(url) if url else "NAME::" + re.sub(r"\s+", "", name)
        if key in out:
            continue
        emp   = {lab: (cv(r[i]) if len(r) > i else "") for i, lab in MAIL_EMP_COLS.items()}
        owner = {lab: (cv(r[i]) if len(r) > i else "") for i, lab in MAIL_OWNER_COLS.items()}
        full  = [r[i] if i < len(r) else None for i in range(34)]   # 신규행 append용 원본 34열
        out[key] = {"name": name, "url": url, "row": rownum, "emp": emp, "owner": owner, "full": full}
        order.append(key)
    wb.close()
    return out, order

def load_inf(path):
    """반환: {key: {'name','phone','col','cells': {label:val}}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[INF_SHEET]
    rows = list(ws.iter_rows(min_row=1, max_row=46, values_only=True))
    labels = {r: cv(rows[r-1][0]) for r in INF_LABEL_ROWS if r-1 < len(rows)}
    name_row = rows[INF_ROW_NAME-1]
    out, order = {}, []
    for col in range(2, len(name_row)):
        name = cv(name_row[col])
        if not name:
            continue
        phone = norm_phone(rows[INF_ROW_PHONE-1][col]) if INF_ROW_PHONE-1 < len(rows) else ""
        key = phone if phone else "NAME::" + re.sub(r"\s+", "", name)
        cells = {labels.get(r, str(r)): cv(rows[r-1][col]) for r in INF_LABEL_ROWS if r-1 < len(rows)}
        out[key] = {"name": name, "phone": phone, "col": col, "cells": cells}
        order.append(key)
    wb.close()
    return out, order, labels

# ---------- 스냅샷 ----------
def snap_path(kind):
    return os.path.join(SNAP_DIR, "snapshot_%s.json" % kind)

def save_snapshot(stamp):
    """현재 로컬+공유본 상태를 스냅샷으로 저장. stamp=호출측이 넘긴 시각 문자열."""
    os.makedirs(SNAP_DIR, exist_ok=True)
    local = latest_local()
    lm, _ = load_mail(local)
    li, _, _ = load_inf(local)
    sm, _ = load_mail(SHARED_FILE)
    si, _, _ = load_inf(SHARED_FILE)
    snap = {
        "stamp": stamp,
        "local_file": os.path.basename(local),
        "shared_file": os.path.basename(SHARED_FILE),
        # 메일발송현황: urlkey -> emp 컬럼 dict (+존재). 형 열은 로컬 권위라 미저장.
        "mail": {k: v["emp"] for k, v in lm.items()},
        "mail_shared": {k: v["emp"] for k, v in sm.items()},
        # 인플관리: key -> cells dict
        "inf": {k: v["cells"] for k, v in li.items()},
        "inf_shared": {k: v["cells"] for k, v in si.items()},
        "inf_names": {k: v["name"] for k, v in li.items()},
    }
    p = snap_path("base")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False)
    return p, len(snap["mail"]), len(snap["inf"]), len(snap["mail_shared"]), len(snap["inf_shared"])

def load_snapshot():
    p = snap_path("base")
    if not os.path.exists(p):
        return None
    with open(p, encoding="utf-8") as f:
        return json.load(f)

if __name__ == "__main__":
    import sys
    stamp = sys.argv[1] if len(sys.argv) > 1 else "manual"
    p, mn, infn, smn, sin = save_snapshot(stamp)
    print("스냅샷 저장:", p)
    print("  로컬   메일 %d행 / 인플관리 %d명" % (mn, infn))
    print("  공유본 메일 %d행 / 인플관리 %d명" % (smn, sin))
